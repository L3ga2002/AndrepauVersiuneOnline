from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from datetime import datetime, timezone
import io

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from database import db
from auth import require_admin

router = APIRouter()


@router.get("/backup")
async def create_backup(user: dict = Depends(require_admin)):
    backup_data = {
        "created_at": datetime.now(timezone.utc).isoformat(),
        "products": await db.products.find({}, {"_id": 0}).to_list(100000),
        "suppliers": await db.suppliers.find({}, {"_id": 0}).to_list(10000),
        "sales": await db.sales.find({}, {"_id": 0}).to_list(100000),
        "nirs": await db.nirs.find({}, {"_id": 0}).to_list(10000),
        "users": await db.users.find({}, {"_id": 0, "password": 0}).to_list(1000)
    }
    return backup_data


@router.get("/backup/products-csv")
async def export_products_csv(user: dict = Depends(require_admin)):
    products = await db.products.find({}, {"_id": 0}).to_list(100000)

    headers = ["Denumire", "Categorie", "Cod Bare", "Pret Achizitie", "Pret Vanzare", "TVA %", "Unitate", "Stoc", "Stoc Minim"]
    rows = []
    for p in products:
        rows.append([
            p.get("nume", ""),
            p.get("categorie", ""),
            p.get("cod_bare", "") or "",
            str(p.get("pret_achizitie", 0)),
            str(p.get("pret_vanzare", 0)),
            str(p.get("tva", 21)),
            p.get("unitate", "buc"),
            str(p.get("stoc", 0)),
            str(p.get("stoc_minim", 5))
        ])

    return {
        "headers": headers,
        "rows": rows,
        "total": len(rows)
    }


@router.get("/products/export/xls")
async def export_products_xls(user: dict = Depends(require_admin)):
    products = await db.products.find({}, {"_id": 0}).to_list(100000)

    wb = Workbook()
    ws = wb.active
    ws.title = "Produse ANDREPAU"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="F59E0B", end_color="F59E0B", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    headers = ["Nr.", "Denumire", "Categorie", "Cod Bare", "Preț Achiziție", "Preț Vânzare", "TVA %", "Unitate", "Stoc", "Stoc Minim"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    for idx, p in enumerate(products, 1):
        row = [
            idx,
            p.get("nume", ""),
            p.get("categorie", ""),
            p.get("cod_bare", "") or "",
            p.get("pret_achizitie", 0),
            p.get("pret_vanzare", 0),
            p.get("tva", 21),
            p.get("unitate", "buc"),
            p.get("stoc", 0),
            p.get("stoc_minim", 5)
        ]
        for col, value in enumerate(row, 1):
            cell = ws.cell(row=idx + 1, column=col, value=value)
            cell.border = thin_border
            if col in [5, 6]:
                cell.number_format = '#,##0.00 "RON"'
            elif col in [9, 10]:
                cell.number_format = '#,##0.00'

    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 12
    ws.column_dimensions['J'].width = 12

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"produse_andrepau_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
