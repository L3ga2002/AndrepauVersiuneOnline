from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from typing import Optional
import csv
import io
import uuid
from datetime import datetime, timezone
import logging

from database import db
from auth import get_current_user, require_admin
from models import ProductCreate, ProductUpdate, ProductResponse
from utils import parse_number

router = APIRouter()
logger = logging.getLogger(__name__)


@router.post("/products", response_model=ProductResponse)
async def create_product(product: ProductCreate, user: dict = Depends(require_admin)):
    product_doc = {
        "id": str(uuid.uuid4()),
        **product.model_dump(),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.products.insert_one(product_doc)
    return ProductResponse(**{k: v for k, v in product_doc.items() if k != "_id"})


@router.get("/products")
async def get_products(
    search: Optional[str] = None,
    price: Optional[str] = None,
    categorie: Optional[str] = None,
    low_stock: Optional[bool] = None,
    page: int = 1,
    limit: int = 50,
    user: dict = Depends(get_current_user)
):
    query = {}
    conditions = []

    if search:
        search = search.strip()
        words = search.split()
        text_words = [w for w in words if len(w) >= 1]

        if len(text_words) > 1:
            for w in text_words:
                if len(w) >= 2:
                    conditions.append({"nume": {"$regex": w, "$options": "i"}})
        elif text_words:
            conditions.append({"$or": [
                {"nume": {"$regex": search, "$options": "i"}},
                {"cod_bare": {"$regex": search, "$options": "i"}}
            ]})

    if price:
        try:
            price_val = float(price.strip().replace(',', '.'))
            conditions.append({"$or": [
                {"pret_vanzare": {"$gte": round(price_val - 0.10, 2), "$lte": round(price_val + 0.10, 2)}},
                {"pret_achizitie": {"$gte": round(price_val - 0.10, 2), "$lte": round(price_val + 0.10, 2)}},
            ]})
        except ValueError:
            pass

    if conditions:
        query["$and"] = conditions
    if categorie:
        query["categorie"] = categorie
    if low_stock:
        query["$expr"] = {"$lte": ["$stoc", "$stoc_minim"]}

    total = await db.products.count_documents(query)
    skip = (page - 1) * limit
    products = await db.products.find(query, {"_id": 0}).skip(skip).limit(limit).to_list(limit)

    return {
        "products": [ProductResponse(**p) for p in products],
        "total": total,
        "page": page,
        "limit": limit,
        "pages": (total + limit - 1) // limit
    }


@router.get("/products/csv-template")
async def get_csv_template(user: dict = Depends(require_admin)):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Import Produse"

    headers = ["Denumire", "Categorie", "Cod Bare", "Preț Achiziție", "Preț Vânzare", "TVA %", "Unitate", "Stoc", "Stoc Minim"]
    header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style='thin', color='555555'),
        right=Side(style='thin', color='555555'),
        top=Side(style='thin', color='555555'),
        bottom=Side(style='thin', color='555555')
    )

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    widths = [35, 25, 18, 16, 16, 10, 12, 10, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    example_data = [
        ["Ciment Romcim 40kg", "Materiale Construcții", "5941234567890", 22.50, 29.90, 19, "sac", 100, 10],
        ["Fier beton 12mm PC52", "Materiale Construcții", "5941234567891", 38.00, 49.50, 19, "buc", 200, 20],
        ["Vopsea lavabilă albă 15L", "Vopsele", "", 65.00, 85.00, 19, "buc", 30, 5],
        ["Adeziv gresie 25kg", "Adezivi", "5941234567893", 22.00, 28.90, 19, "sac", 50, 10],
        ["Polistiren expandat 10cm", "Izolații", "", 9.50, 12.50, 19, "mp", 100, 20],
    ]

    data_font = Font(name="Calibri", size=11)
    price_format = '#,##0.00'
    yellow_fill = PatternFill(start_color="FFF9E6", end_color="FFF9E6", fill_type="solid")

    for row_idx, row_data in enumerate(example_data, 2):
        bg = yellow_fill if row_idx % 2 == 0 else None
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left' if col_idx <= 3 else 'center')
            if bg:
                cell.fill = bg
            if col_idx in [4, 5]:
                cell.number_format = price_format
                cell.alignment = Alignment(horizontal='right')

    ws.cell(row=8, column=1, value="INSTRUCȚIUNI:").font = Font(bold=True, color="FF6600", size=10)
    ws.cell(row=9, column=1, value="- Coloana 'Denumire' și 'Preț Vânzare' sunt OBLIGATORII").font = Font(size=9, color="666666")
    ws.cell(row=10, column=1, value="- Completați 'Cod Bare' pentru potrivire automată la import").font = Font(size=9, color="666666")
    ws.cell(row=11, column=1, value="- Unitate: buc, sac, kg, metru, litru, rola").font = Font(size=9, color="666666")
    ws.cell(row=12, column=1, value="- Ștergeți exemplele de mai sus și completați cu produsele dvs.").font = Font(size=9, color="666666")
    ws.cell(row=13, column=1, value="- Salvați ca .xlsx sau .csv apoi importați din aplicație").font = Font(size=9, color="666666")

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = 'A2'

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=template_import_produse.xlsx"}
    )


@router.post("/products/import-csv")
async def import_products_file(file: UploadFile = File(...), user: dict = Depends(require_admin)):
    fname = file.filename.lower()
    if not (fname.endswith('.csv') or fname.endswith('.xlsx') or fname.endswith('.xls')):
        raise HTTPException(status_code=400, detail="Fișierul trebuie să fie .xlsx sau .csv")

    content = await file.read()
    if len(content) > 10 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Fișierul este prea mare (max 10MB)")

    rows = []
    if fname.endswith('.xlsx'):
        from openpyxl import load_workbook
        try:
            wb = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
            ws = wb.active
            for row in ws.iter_rows(values_only=True):
                str_row = [str(c).strip() if c is not None else '' for c in row]
                if any(c for c in str_row):
                    rows.append(str_row)
            wb.close()
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Eroare la citire Excel: {str(e)}")
    elif fname.endswith('.xls'):
        import xlrd
        try:
            wb = xlrd.open_workbook(file_contents=content)
            ws = wb.sheet_by_index(0)
            for r in range(ws.nrows):
                str_row = [str(ws.cell_value(r, c)).strip() if ws.cell_value(r, c) != '' else '' for c in range(ws.ncols)]
                if any(c for c in str_row):
                    rows.append(str_row)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Eroare la citire Excel .xls: {str(e)}")
    else:
        text = None
        for encoding in ['utf-8-sig', 'utf-8', 'latin-1', 'cp1252']:
            try:
                text = content.decode(encoding)
                break
            except (UnicodeDecodeError, ValueError):
                continue
        if text is None:
            raise HTTPException(status_code=400, detail="Nu am putut citi fișierul. Verificați codificarea.")
        first_line = text.split('\n')[0]
        delimiter = ';' if ';' in first_line else ','
        reader = csv.reader(io.StringIO(text), delimiter=delimiter)
        rows = list(reader)

    if len(rows) < 2:
        raise HTTPException(status_code=400, detail="Fișierul trebuie să aibă cel puțin un rând de date (+ header)")

    def _strip_diacritics(s):
        replacements = {'ă': 'a', 'â': 'a', 'î': 'i', 'ș': 's', 'ț': 't',
                        'Ă': 'A', 'Â': 'A', 'Î': 'I', 'Ș': 'S', 'Ț': 'T',
                        'ş': 's', 'ţ': 't'}
        for old, new in replacements.items():
            s = s.replace(old, new)
        return s

    header = [_strip_diacritics(h.strip().lower()) for h in rows[0]]

    col_map = {}
    name_variants = {"denumire", "nume", "produs", "name", "denumire produs"}
    cat_variants = {"categorie", "category", "cat"}
    barcode_variants = {"cod bare", "cod_bare", "barcode", "ean", "cod"}
    buy_variants = {"pret achizitie", "pret_achizitie", "pret cumparare", "achizitie", "pret achizitie (ron)", "pretachizitie"}
    sell_variants = {"pret vanzare", "pret_vanzare", "vanzare", "pret", "pret vanzare (ron)", "pretvanzare"}
    tva_variants = {"tva", "tva %", "tva%", "cota tva", "cotatva", "cota_tva"}
    unit_variants = {"unitate", "um", "unit", "unitate masura"}
    stock_variants = {"stoc", "stock", "cantitate", "stoc disponibil", "stocinitial", "stoc initial", "stoc_initial"}
    min_stock_variants = {"stoc minim", "stoc_minim", "min stock", "stoc min", "stoc minim alerta"}

    for idx, h in enumerate(header):
        if h in name_variants:
            col_map["nume"] = idx
        elif h in cat_variants:
            col_map["categorie"] = idx
        elif h in barcode_variants:
            col_map["cod_bare"] = idx
        elif h in buy_variants:
            col_map["pret_achizitie"] = idx
        elif h in sell_variants:
            col_map["pret_vanzare"] = idx
        elif h in tva_variants:
            col_map["tva"] = idx
        elif h in unit_variants:
            col_map["unitate"] = idx
        elif h in stock_variants:
            col_map["stoc"] = idx
        elif h in min_stock_variants:
            col_map["stoc_minim"] = idx

    if "nume" not in col_map:
        raise HTTPException(status_code=400, detail="Coloana 'Denumire' este obligatorie în CSV")
    if "pret_vanzare" not in col_map:
        raise HTTPException(status_code=400, detail="Coloana 'Pret Vanzare' este obligatorie în CSV")

    parsed = []
    errors = []
    existing_products = await db.products.find({}, {"_id": 0, "id": 1, "nume": 1, "cod_bare": 1}).to_list(100000)
    existing_by_name = {p["nume"].lower(): p for p in existing_products}
    existing_by_barcode = {p["cod_bare"]: p for p in existing_products if p.get("cod_bare")}

    for row_idx, row in enumerate(rows[1:], start=2):
        if not row or all(c.strip() == '' for c in row):
            continue

        try:
            name = row[col_map["nume"]].strip() if col_map.get("nume") is not None and col_map["nume"] < len(row) else ""
            if not name:
                errors.append(f"Rândul {row_idx}: Denumire lipsă")
                continue

            sell_price_str = row[col_map["pret_vanzare"]].strip() if col_map.get("pret_vanzare") is not None and col_map["pret_vanzare"] < len(row) else "0"
            sell_price = parse_number(sell_price_str)
            if sell_price < 0:
                errors.append(f"Rândul {row_idx}: Preț vânzare negativ pentru '{name}'")
                continue

            barcode = row[col_map["cod_bare"]].strip() if col_map.get("cod_bare") is not None and col_map["cod_bare"] < len(row) else ""
            category = row[col_map["categorie"]].strip() if col_map.get("categorie") is not None and col_map["categorie"] < len(row) else "Necategorisit"
            buy_price = parse_number(row[col_map["pret_achizitie"]].strip()) if col_map.get("pret_achizitie") is not None and col_map["pret_achizitie"] < len(row) else 0
            tva = parse_number(row[col_map["tva"]].strip()) if col_map.get("tva") is not None and col_map["tva"] < len(row) else 19
            unit = row[col_map["unitate"]].strip() if col_map.get("unitate") is not None and col_map["unitate"] < len(row) else "buc"
            unit_normalize = {"m": "metru", "l": "litru", "metri": "metru", "litri": "litru", "bucati": "buc", "saci": "sac", "role": "rola"}
            unit = unit_normalize.get(unit.lower(), unit.lower()) if unit else "buc"
            stock = parse_number(row[col_map["stoc"]].strip()) if col_map.get("stoc") is not None and col_map["stoc"] < len(row) else 0
            min_stock = parse_number(row[col_map["stoc_minim"]].strip()) if col_map.get("stoc_minim") is not None and col_map["stoc_minim"] < len(row) else 5

            existing = None
            action = "create"
            if barcode and barcode in existing_by_barcode:
                existing = existing_by_barcode[barcode]
                action = "update"
            elif name.lower() in existing_by_name:
                existing = existing_by_name[name.lower()]
                action = "update"

            parsed.append({
                "row": row_idx,
                "nume": name,
                "categorie": category,
                "cod_bare": barcode,
                "pret_achizitie": buy_price,
                "pret_vanzare": sell_price,
                "tva": tva,
                "unitate": unit if unit in ["buc", "sac", "kg", "metru", "litru", "rola"] else "buc",
                "stoc": stock,
                "stoc_minim": min_stock,
                "action": action,
                "existing_id": existing["id"] if existing else None,
                "existing_name": existing["nume"] if existing else None,
            })

        except (IndexError, ValueError) as e:
            errors.append(f"Rândul {row_idx}: Eroare la parsare - {str(e)}")

    return {
        "items": parsed,
        "errors": errors,
        "total_parsed": len(parsed),
        "total_create": len([p for p in parsed if p["action"] == "create"]),
        "total_update": len([p for p in parsed if p["action"] == "update"]),
        "columns_found": list(col_map.keys())
    }


@router.post("/products/import-csv/confirm")
async def confirm_import_products_csv(data: dict, user: dict = Depends(require_admin)):
    items = data.get("items", [])
    if not items:
        raise HTTPException(status_code=400, detail="Nu sunt produse de importat")

    from pymongo import UpdateOne, InsertOne

    created = 0
    updated = 0
    errors = []
    operations = []

    for item in items:
        try:
            if item.get("action") == "update" and item.get("existing_id"):
                update_fields = {
                    "categorie": item["categorie"],
                    "pret_achizitie": item["pret_achizitie"],
                    "pret_vanzare": item["pret_vanzare"],
                    "tva": item["tva"],
                    "unitate": item["unitate"],
                    "stoc": item["stoc"],
                    "stoc_minim": item["stoc_minim"],
                }
                if item.get("cod_bare"):
                    update_fields["cod_bare"] = item["cod_bare"]
                operations.append(UpdateOne(
                    {"id": item["existing_id"]},
                    {"$set": update_fields}
                ))
                updated += 1
            else:
                product_doc = {
                    "id": str(uuid.uuid4()),
                    "nume": item["nume"],
                    "categorie": item["categorie"],
                    "cod_bare": item.get("cod_bare", ""),
                    "pret_achizitie": item["pret_achizitie"],
                    "pret_vanzare": item["pret_vanzare"],
                    "tva": item["tva"],
                    "unitate": item["unitate"],
                    "stoc": item["stoc"],
                    "stoc_minim": item["stoc_minim"],
                    "descriere": "",
                    "furnizor_id": None,
                    "created_at": datetime.now(timezone.utc).isoformat()
                }
                operations.append(InsertOne(product_doc))
                created += 1
        except Exception as e:
            errors.append(f"Eroare la '{item.get('nume', '?')}': {str(e)}")

    if operations:
        BATCH = 500
        for i in range(0, len(operations), BATCH):
            batch = operations[i:i+BATCH]
            try:
                await db.products.bulk_write(batch, ordered=False)
            except Exception as e:
                errors.append(f"Eroare bulk batch {i//BATCH+1}: {str(e)}")

    return {
        "created": created,
        "updated": updated,
        "errors": errors,
        "total": created + updated
    }


@router.get("/products/barcode/{barcode}", response_model=ProductResponse)
async def get_product_by_barcode(barcode: str, user: dict = Depends(get_current_user)):
    product = await db.products.find_one({"cod_bare": barcode}, {"_id": 0})
    if not product:
        raise HTTPException(status_code=404, detail="Produs negăsit")
    return ProductResponse(**product)


@router.get("/products/{product_id}", response_model=ProductResponse)
async def get_product(product_id: str, user: dict = Depends(get_current_user)):
    product = await db.products.find_one({"id": product_id}, {"_id": 0})
    if not product:
        raise HTTPException(status_code=404, detail="Produs negăsit")
    return ProductResponse(**product)


@router.put("/products/{product_id}", response_model=ProductResponse)
async def update_product(product_id: str, product: ProductUpdate, user: dict = Depends(require_admin)):
    update_data = {k: v for k, v in product.model_dump().items() if v is not None}
    if not update_data:
        raise HTTPException(status_code=400, detail="Nicio modificare")

    result = await db.products.update_one({"id": product_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Produs negăsit")

    updated = await db.products.find_one({"id": product_id}, {"_id": 0})
    return ProductResponse(**updated)


@router.delete("/products/{product_id}")
async def delete_product(product_id: str, user: dict = Depends(require_admin)):
    result = await db.products.delete_one({"id": product_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Produs negăsit")
    return {"message": "Produs șters"}


@router.delete("/products-all/delete")
async def delete_all_products(user: dict = Depends(require_admin)):
    result = await db.products.delete_many({})
    logger.warning(f"ALL PRODUCTS DELETED by {user['username']} - {result.deleted_count} products removed")
    return {"message": f"{result.deleted_count} produse șterse", "deleted_count": result.deleted_count}


@router.put("/products-all/bulk-tva")
async def bulk_update_tva(data: dict, user: dict = Depends(require_admin)):
    new_tva = data.get("tva")
    if new_tva is None or not isinstance(new_tva, (int, float)) or new_tva < 0 or new_tva > 100:
        raise HTTPException(status_code=400, detail="Cota TVA trebuie să fie între 0 și 100")
    result = await db.products.update_many({}, {"$set": {"tva": float(new_tva)}})
    logger.info(f"TVA updated to {new_tva}% for {result.modified_count} products by {user['username']}")
    return {"message": f"TVA actualizat la {new_tva}% pentru {result.modified_count} produse", "modified_count": result.modified_count}


@router.post("/products/bulk-barcode")
async def bulk_update_barcodes(data: dict, user: dict = Depends(require_admin)):
    """Salveaza codurile de bare pe produse. Daca un cod de bare introdus exista DEJA pe
    alt produs, face MERGE: transfera stocul pe produsul existent si sterge produsul nou (duplicat).
    Astfel evitam dubluri cand la import NIR PDF se creeaza produs nou cu denumire usor diferita."""
    updates = data.get("updates", [])
    if not updates:
        raise HTTPException(status_code=400, detail="Nu sunt actualizări")

    updated = 0
    merged = 0
    merge_details = []
    now_iso = datetime.now(timezone.utc).isoformat()

    for upd in updates:
        product_id = upd.get("product_id")
        barcode = upd.get("cod_bare", "").strip()
        if not product_id or not barcode:
            continue

        # Check if barcode already exists on ANOTHER product
        existing_with_barcode = await db.products.find_one(
            {"cod_bare": barcode, "id": {"$ne": product_id}},
            {"_id": 0}
        )

        if existing_with_barcode:
            # MERGE: take stock from current product, add to existing, delete current
            current = await db.products.find_one({"id": product_id}, {"_id": 0})
            if current:
                stoc_transfer = current.get("stoc", 0)
                # Actualizeaza pret_vanzare si stoc pe produsul existent (pastreaza denumire!)
                await db.products.update_one(
                    {"id": existing_with_barcode["id"]},
                    {
                        "$inc": {"stoc": stoc_transfer},
                        "$set": {
                            "pret_vanzare": current.get("pret_vanzare", existing_with_barcode.get("pret_vanzare", 0)),
                            "updated_at": now_iso
                        }
                    }
                )
                # Sterge produsul duplicat
                await db.products.delete_one({"id": product_id})
                merged += 1
                merge_details.append({
                    "deleted_product": current.get("nume", ""),
                    "merged_into": existing_with_barcode.get("nume", ""),
                    "stoc_transferat": stoc_transfer
                })
                logger.info(
                    f"[BARCODE-MERGE] '{current.get('nume','')}' -> '{existing_with_barcode.get('nume','')}' "
                    f"(cod_bare={barcode}, stoc_transferat={stoc_transfer})"
                )
        else:
            # Normal update - set barcode on product
            result = await db.products.update_one(
                {"id": product_id},
                {"$set": {"cod_bare": barcode, "updated_at": now_iso}}
            )
            if result.modified_count > 0:
                updated += 1

    return {
        "updated": updated,
        "merged": merged,
        "merge_details": merge_details,
        "total": len(updates)
    }


@router.post("/products/update-tva")
async def update_products_tva(data: dict, user: dict = Depends(get_current_user)):
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Doar administratorul poate modifica TVA")

    old_tva = data.get("old_tva", 19.0)
    new_tva = data.get("new_tva", 21.0)

    result = await db.products.update_many(
        {"tva": old_tva},
        {"$set": {"tva": new_tva}}
    )

    return {
        "message": f"TVA actualizat de la {old_tva}% la {new_tva}%",
        "products_updated": result.modified_count
    }


@router.get("/categories")
async def get_categories(user: dict = Depends(get_current_user)):
    categories = await db.products.distinct("categorie")
    return categories
