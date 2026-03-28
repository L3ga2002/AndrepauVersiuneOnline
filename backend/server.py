from fastapi import FastAPI, APIRouter, HTTPException, Depends, status, UploadFile, File
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.responses import StreamingResponse, FileResponse
from fastapi.staticfiles import StaticFiles
import csv
import re
import fitz  # PyMuPDF
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict
from typing import List, Optional
import uuid
from datetime import datetime, timezone, timedelta
import jwt
import bcrypt
from bson import ObjectId
import json
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# JWT Config
JWT_SECRET = os.environ.get('JWT_SECRET', 'andrepau-secret-key-2024')
JWT_ALGORITHM = "HS256"
JWT_EXPIRATION_HOURS = 24

# Create the main app
app = FastAPI(title="ANDREPAU POS API")

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

security = HTTPBearer()

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# ==================== MODELS ====================

class UserCreate(BaseModel):
    username: str
    password: str
    full_name: str
    role: str = "casier"  # admin or casier

class UserLogin(BaseModel):
    username: str
    password: str

class UserResponse(BaseModel):
    id: str
    username: str
    full_name: str
    role: str
    created_at: str

class ProductCreate(BaseModel):
    nume: str
    categorie: str
    furnizor_id: Optional[str] = None
    cod_bare: Optional[str] = None
    pret_achizitie: float = 0
    pret_vanzare: float
    tva: float = 21.0
    unitate: str = "buc"  # buc, sac, kg, metru, litru, rola
    stoc: float = 0
    stoc_minim: float = 5
    descriere: Optional[str] = None

class ProductUpdate(BaseModel):
    nume: Optional[str] = None
    categorie: Optional[str] = None
    furnizor_id: Optional[str] = None
    cod_bare: Optional[str] = None
    pret_achizitie: Optional[float] = None
    pret_vanzare: Optional[float] = None
    tva: Optional[float] = None
    unitate: Optional[str] = None
    stoc: Optional[float] = None
    stoc_minim: Optional[float] = None
    descriere: Optional[str] = None

class ProductResponse(BaseModel):
    id: str
    nume: str
    categorie: str
    furnizor_id: Optional[str] = None
    cod_bare: Optional[str] = None
    pret_achizitie: float
    pret_vanzare: float
    tva: float
    unitate: str
    stoc: float
    stoc_minim: float
    descriere: Optional[str] = None
    created_at: str

class SupplierCreate(BaseModel):
    nume: str
    telefon: Optional[str] = None
    email: Optional[str] = None
    adresa: Optional[str] = None

class SupplierResponse(BaseModel):
    id: str
    nume: str
    telefon: Optional[str] = None
    email: Optional[str] = None
    adresa: Optional[str] = None
    created_at: str

class SaleItem(BaseModel):
    product_id: str
    nume: str
    cantitate: float
    pret_unitar: float
    tva: float

class SaleCreate(BaseModel):
    items: List[SaleItem]
    subtotal: float
    tva_total: float
    total: float
    discount_percent: float = 0
    metoda_plata: str  # numerar, card, combinat
    suma_numerar: float = 0
    suma_card: float = 0
    casier_id: str
    transaction_id: Optional[str] = None
    fiscal_number: Optional[str] = None
    fiscal_status: str = "none"  # none, printed, cancelled

class SaleResponse(BaseModel):
    id: str
    numar_bon: str
    items: List[SaleItem]
    subtotal: float
    tva_total: float
    total: float
    discount_percent: float
    metoda_plata: str
    suma_numerar: float
    suma_card: float
    casier_id: str
    casier_nume: str
    created_at: str
    transaction_id: Optional[str] = None
    fiscal_number: Optional[str] = None
    fiscal_status: str = "none"

class NIRItem(BaseModel):
    product_id: str
    nume: str
    cantitate: float
    pret_achizitie: float

class NIRCreate(BaseModel):
    furnizor_id: str
    numar_factura: str
    items: List[NIRItem]
    total: float

class NIRResponse(BaseModel):
    id: str
    numar_nir: str
    furnizor_id: str
    furnizor_nume: str
    numar_factura: str
    items: List[NIRItem]
    total: float
    created_at: str

# ==================== AUTH HELPERS ====================

def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')

def verify_password(password: str, hashed: str) -> bool:
    return bcrypt.checkpw(password.encode('utf-8'), hashed.encode('utf-8'))

def create_token(user_id: str, role: str) -> str:
    payload = {
        "user_id": user_id,
        "role": role,
        "exp": datetime.now(timezone.utc) + timedelta(hours=JWT_EXPIRATION_HOURS)
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    try:
        payload = jwt.decode(credentials.credentials, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        user = await db.users.find_one({"id": payload["user_id"]}, {"_id": 0})
        if not user:
            raise HTTPException(status_code=401, detail="Utilizator negăsit")
        return user
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expirat")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Token invalid")

async def require_admin(user: dict = Depends(get_current_user)):
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Acces permis doar pentru admin")
    return user

# ==================== AUTH ROUTES ====================

@api_router.post("/auth/register", response_model=UserResponse)
async def register(user: UserCreate):
    existing = await db.users.find_one({"username": user.username})
    if existing:
        raise HTTPException(status_code=400, detail="Username-ul există deja")
    
    user_doc = {
        "id": str(uuid.uuid4()),
        "username": user.username,
        "password": hash_password(user.password),
        "full_name": user.full_name,
        "role": user.role,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.users.insert_one(user_doc)
    
    return UserResponse(
        id=user_doc["id"],
        username=user_doc["username"],
        full_name=user_doc["full_name"],
        role=user_doc["role"],
        created_at=user_doc["created_at"]
    )

@api_router.post("/auth/login")
async def login(credentials: UserLogin):
    user = await db.users.find_one({"username": credentials.username}, {"_id": 0})
    if not user or not verify_password(credentials.password, user["password"]):
        raise HTTPException(status_code=401, detail="Credențiale invalide")
    
    token = create_token(user["id"], user["role"])
    return {
        "token": token,
        "user": {
            "id": user["id"],
            "username": user["username"],
            "full_name": user["full_name"],
            "role": user["role"]
        }
    }

@api_router.get("/auth/me", response_model=UserResponse)
async def get_me(user: dict = Depends(get_current_user)):
    return UserResponse(
        id=user["id"],
        username=user["username"],
        full_name=user["full_name"],
        role=user["role"],
        created_at=user["created_at"]
    )

@api_router.get("/users", response_model=List[UserResponse])
async def get_users(user: dict = Depends(require_admin)):
    users = await db.users.find({}, {"_id": 0, "password": 0}).to_list(1000)
    return [UserResponse(**u) for u in users]

@api_router.delete("/users/{user_id}")
async def delete_user(user_id: str, admin: dict = Depends(require_admin)):
    if admin["id"] == user_id:
        raise HTTPException(status_code=400, detail="Nu vă puteți șterge propriu cont")
    result = await db.users.delete_one({"id": user_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Utilizator negăsit")
    return {"message": "Utilizator șters"}

# ==================== PRODUCT ROUTES ====================

@api_router.post("/products", response_model=ProductResponse)
async def create_product(product: ProductCreate, user: dict = Depends(require_admin)):
    product_doc = {
        "id": str(uuid.uuid4()),
        **product.model_dump(),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.products.insert_one(product_doc)
    return ProductResponse(**{k: v for k, v in product_doc.items() if k != "_id"})

@api_router.get("/products")
async def get_products(
    search: Optional[str] = None,
    price: Optional[str] = None,
    categorie: Optional[str] = None,
    low_stock: Optional[bool] = None,
    page: int = 1,
    limit: int = 50,
    user: dict = Depends(get_current_user)
):
    query = {}
    conditions = []

    if search:
        search = search.strip()
        words = search.split()
        text_words = [w for w in words if len(w) >= 1]

        if len(text_words) > 1:
            # Multi-word: each word must appear in name
            for w in text_words:
                if len(w) >= 2:
                    conditions.append({"nume": {"$regex": w, "$options": "i"}})
        elif text_words:
            # Single word: name or barcode match
            conditions.append({"$or": [
                {"nume": {"$regex": search, "$options": "i"}},
                {"cod_bare": {"$regex": search, "$options": "i"}}
            ]})

    if price:
        try:
            price_val = float(price.strip().replace(',', '.'))
            conditions.append({"$or": [
                {"pret_vanzare": {"$gte": price_val - 2, "$lte": price_val + 2}},
                {"pret_achizitie": {"$gte": price_val - 2, "$lte": price_val + 2}},
            ]})
        except ValueError:
            pass

    if conditions:
        query["$and"] = conditions
    if categorie:
        query["categorie"] = categorie
    if low_stock:
        query["$expr"] = {"$lte": ["$stoc", "$stoc_minim"]}
    
    # Get total count
    total = await db.products.count_documents(query)
    
    # Get paginated results
    skip = (page - 1) * limit
    products = await db.products.find(query, {"_id": 0}).skip(skip).limit(limit).to_list(limit)
    
    return {
        "products": [ProductResponse(**p) for p in products],
        "total": total,
        "page": page,
        "limit": limit,
        "pages": (total + limit - 1) // limit
    }


# ==================== PRODUCTS CSV IMPORT ====================

@api_router.get("/products/csv-template")
async def get_csv_template(user: dict = Depends(require_admin)):
    """Download Excel template for product import"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Import Produse"

    # Headers
    headers = ["Denumire", "Categorie", "Cod Bare", "Preț Achiziție", "Preț Vânzare", "TVA %", "Unitate", "Stoc", "Stoc Minim"]
    header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style='thin', color='555555'),
        right=Side(style='thin', color='555555'),
        top=Side(style='thin', color='555555'),
        bottom=Side(style='thin', color='555555')
    )

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # Column widths
    widths = [35, 25, 18, 16, 16, 10, 12, 10, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    # Example data
    example_data = [
        ["Ciment Romcim 40kg", "Materiale Construcții", "5941234567890", 22.50, 29.90, 19, "sac", 100, 10],
        ["Fier beton 12mm PC52", "Materiale Construcții", "5941234567891", 38.00, 49.50, 19, "buc", 200, 20],
        ["Vopsea lavabilă albă 15L", "Vopsele", "", 65.00, 85.00, 19, "buc", 30, 5],
        ["Adeziv gresie 25kg", "Adezivi", "5941234567893", 22.00, 28.90, 19, "sac", 50, 10],
        ["Polistiren expandat 10cm", "Izolații", "", 9.50, 12.50, 19, "mp", 100, 20],
    ]

    data_font = Font(name="Calibri", size=11)
    price_format = '#,##0.00'
    yellow_fill = PatternFill(start_color="FFF9E6", end_color="FFF9E6", fill_type="solid")

    for row_idx, row_data in enumerate(example_data, 2):
        bg = yellow_fill if row_idx % 2 == 0 else None
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left' if col_idx <= 3 else 'center')
            if bg:
                cell.fill = bg
            if col_idx in [4, 5]:
                cell.number_format = price_format
                cell.alignment = Alignment(horizontal='right')

    # Instructions row
    ws.cell(row=8, column=1, value="INSTRUCȚIUNI:").font = Font(bold=True, color="FF6600", size=10)
    ws.cell(row=9, column=1, value="- Coloana 'Denumire' și 'Preț Vânzare' sunt OBLIGATORII").font = Font(size=9, color="666666")
    ws.cell(row=10, column=1, value="- Completați 'Cod Bare' pentru potrivire automată la import").font = Font(size=9, color="666666")
    ws.cell(row=11, column=1, value="- Unitate: buc, sac, kg, metru, litru, rola").font = Font(size=9, color="666666")
    ws.cell(row=12, column=1, value="- Ștergeți exemplele de mai sus și completați cu produsele dvs.").font = Font(size=9, color="666666")
    ws.cell(row=13, column=1, value="- Salvați ca .xlsx sau .csv apoi importați din aplicație").font = Font(size=9, color="666666")

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = 'A2'

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=template_import_produse.xlsx"}
    )


@api_router.post("/products/import-csv")
async def import_products_file(file: UploadFile = File(...), user: dict = Depends(require_admin)):
    """Parse CSV or Excel file and return preview data for confirmation"""
    fname = file.filename.lower()
    if not (fname.endswith('.csv') or fname.endswith('.xlsx') or fname.endswith('.xls')):
        raise HTTPException(status_code=400, detail="Fișierul trebuie să fie .xlsx sau .csv")

    content = await file.read()
    if len(content) > 10 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Fișierul este prea mare (max 10MB)")

    rows = []
    if fname.endswith('.xlsx') or fname.endswith('.xls'):
        # Parse Excel
        from openpyxl import load_workbook
        try:
            wb = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
            ws = wb.active
            for row in ws.iter_rows(values_only=True):
                str_row = [str(c).strip() if c is not None else '' for c in row]
                if any(c for c in str_row):
                    rows.append(str_row)
            wb.close()
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Eroare la citire Excel: {str(e)}")
    else:
        # Parse CSV
        text = None
        for encoding in ['utf-8-sig', 'utf-8', 'latin-1', 'cp1252']:
            try:
                text = content.decode(encoding)
                break
            except (UnicodeDecodeError, ValueError):
                continue
        if text is None:
            raise HTTPException(status_code=400, detail="Nu am putut citi fișierul. Verificați codificarea.")
        first_line = text.split('\n')[0]
        delimiter = ';' if ';' in first_line else ','
        reader = csv.reader(io.StringIO(text), delimiter=delimiter)
        rows = list(reader)

    if len(rows) < 2:
        raise HTTPException(status_code=400, detail="Fișierul trebuie să aibă cel puțin un rând de date (+ header)")

    # Normalize headers - strip diacritics for matching
    def _strip_diacritics(s):
        replacements = {'ă': 'a', 'â': 'a', 'î': 'i', 'ș': 's', 'ț': 't',
                        'Ă': 'A', 'Â': 'A', 'Î': 'I', 'Ș': 'S', 'Ț': 'T',
                        'ş': 's', 'ţ': 't'}
        for old, new in replacements.items():
            s = s.replace(old, new)
        return s

    header = [_strip_diacritics(h.strip().lower()) for h in rows[0]]

    col_map = {}
    name_variants = {"denumire", "nume", "produs", "name", "denumire produs"}
    cat_variants = {"categorie", "category", "cat"}
    barcode_variants = {"cod bare", "cod_bare", "barcode", "ean", "cod"}
    buy_variants = {"pret achizitie", "pret_achizitie", "pret cumparare", "achizitie", "pret achizitie (ron)"}
    sell_variants = {"pret vanzare", "pret_vanzare", "vanzare", "pret", "pret vanzare (ron)"}
    tva_variants = {"tva", "tva %", "tva%", "cota tva"}
    unit_variants = {"unitate", "um", "unit", "unitate masura"}
    stock_variants = {"stoc", "stock", "cantitate", "stoc disponibil"}
    min_stock_variants = {"stoc minim", "stoc_minim", "min stock", "stoc min", "stoc minim alerta"}

    for idx, h in enumerate(header):
        if h in name_variants:
            col_map["nume"] = idx
        elif h in cat_variants:
            col_map["categorie"] = idx
        elif h in barcode_variants:
            col_map["cod_bare"] = idx
        elif h in buy_variants:
            col_map["pret_achizitie"] = idx
        elif h in sell_variants:
            col_map["pret_vanzare"] = idx
        elif h in tva_variants:
            col_map["tva"] = idx
        elif h in unit_variants:
            col_map["unitate"] = idx
        elif h in stock_variants:
            col_map["stoc"] = idx
        elif h in min_stock_variants:
            col_map["stoc_minim"] = idx

    if "nume" not in col_map:
        raise HTTPException(status_code=400, detail="Coloana 'Denumire' este obligatorie în CSV")
    if "pret_vanzare" not in col_map:
        raise HTTPException(status_code=400, detail="Coloana 'Pret Vanzare' este obligatorie în CSV")

    parsed = []
    errors = []
    existing_products = await db.products.find({}, {"_id": 0, "id": 1, "nume": 1, "cod_bare": 1}).to_list(100000)
    existing_by_name = {p["nume"].lower(): p for p in existing_products}
    existing_by_barcode = {p["cod_bare"]: p for p in existing_products if p.get("cod_bare")}

    for row_idx, row in enumerate(rows[1:], start=2):
        if not row or all(c.strip() == '' for c in row):
            continue

        try:
            name = row[col_map["nume"]].strip() if col_map.get("nume") is not None and col_map["nume"] < len(row) else ""
            if not name:
                errors.append(f"Rândul {row_idx}: Denumire lipsă")
                continue

            sell_price_str = row[col_map["pret_vanzare"]].strip() if col_map.get("pret_vanzare") is not None and col_map["pret_vanzare"] < len(row) else "0"
            sell_price = _parse_number(sell_price_str)
            if sell_price <= 0:
                errors.append(f"Rândul {row_idx}: Preț vânzare invalid pentru '{name}'")
                continue

            barcode = row[col_map["cod_bare"]].strip() if col_map.get("cod_bare") is not None and col_map["cod_bare"] < len(row) else ""
            category = row[col_map["categorie"]].strip() if col_map.get("categorie") is not None and col_map["categorie"] < len(row) else "Necategorisit"
            buy_price = _parse_number(row[col_map["pret_achizitie"]].strip()) if col_map.get("pret_achizitie") is not None and col_map["pret_achizitie"] < len(row) else 0
            tva = _parse_number(row[col_map["tva"]].strip()) if col_map.get("tva") is not None and col_map["tva"] < len(row) else 19
            unit = row[col_map["unitate"]].strip() if col_map.get("unitate") is not None and col_map["unitate"] < len(row) else "buc"
            stock = _parse_number(row[col_map["stoc"]].strip()) if col_map.get("stoc") is not None and col_map["stoc"] < len(row) else 0
            min_stock = _parse_number(row[col_map["stoc_minim"]].strip()) if col_map.get("stoc_minim") is not None and col_map["stoc_minim"] < len(row) else 5

            existing = None
            action = "create"
            if barcode and barcode in existing_by_barcode:
                existing = existing_by_barcode[barcode]
                action = "update"
            elif name.lower() in existing_by_name:
                existing = existing_by_name[name.lower()]
                action = "update"

            parsed.append({
                "row": row_idx,
                "nume": name,
                "categorie": category,
                "cod_bare": barcode,
                "pret_achizitie": buy_price,
                "pret_vanzare": sell_price,
                "tva": tva,
                "unitate": unit if unit in ["buc", "sac", "kg", "metru", "litru", "rola"] else "buc",
                "stoc": stock,
                "stoc_minim": min_stock,
                "action": action,
                "existing_id": existing["id"] if existing else None,
                "existing_name": existing["nume"] if existing else None,
            })

        except (IndexError, ValueError) as e:
            errors.append(f"Rândul {row_idx}: Eroare la parsare - {str(e)}")

    return {
        "items": parsed,
        "errors": errors,
        "total_parsed": len(parsed),
        "total_create": len([p for p in parsed if p["action"] == "create"]),
        "total_update": len([p for p in parsed if p["action"] == "update"]),
        "columns_found": list(col_map.keys())
    }


@api_router.post("/products/import-csv/confirm")
async def confirm_import_products_csv(data: dict, user: dict = Depends(require_admin)):
    """Confirm and execute the CSV import"""
    items = data.get("items", [])
    if not items:
        raise HTTPException(status_code=400, detail="Nu sunt produse de importat")

    created = 0
    updated = 0
    errors = []

    for item in items:
        try:
            if item.get("action") == "update" and item.get("existing_id"):
                update_fields = {
                    "categorie": item["categorie"],
                    "pret_achizitie": item["pret_achizitie"],
                    "pret_vanzare": item["pret_vanzare"],
                    "tva": item["tva"],
                    "unitate": item["unitate"],
                    "stoc": item["stoc"],
                    "stoc_minim": item["stoc_minim"],
                }
                if item.get("cod_bare"):
                    update_fields["cod_bare"] = item["cod_bare"]
                await db.products.update_one(
                    {"id": item["existing_id"]},
                    {"$set": update_fields}
                )
                updated += 1
            else:
                product_doc = {
                    "id": str(uuid.uuid4()),
                    "nume": item["nume"],
                    "categorie": item["categorie"],
                    "cod_bare": item.get("cod_bare", ""),
                    "pret_achizitie": item["pret_achizitie"],
                    "pret_vanzare": item["pret_vanzare"],
                    "tva": item["tva"],
                    "unitate": item["unitate"],
                    "stoc": item["stoc"],
                    "stoc_minim": item["stoc_minim"],
                    "descriere": "",
                    "furnizor_id": None,
                    "created_at": datetime.now(timezone.utc).isoformat()
                }
                await db.products.insert_one(product_doc)
                created += 1
        except Exception as e:
            errors.append(f"Eroare la '{item.get('nume', '?')}': {str(e)}")

    return {
        "created": created,
        "updated": updated,
        "errors": errors,
        "total": created + updated
    }


@api_router.get("/products/barcode/{barcode}", response_model=ProductResponse)
async def get_product_by_barcode(barcode: str, user: dict = Depends(get_current_user)):
    product = await db.products.find_one({"cod_bare": barcode}, {"_id": 0})
    if not product:
        raise HTTPException(status_code=404, detail="Produs negăsit")
    return ProductResponse(**product)

@api_router.get("/products/{product_id}", response_model=ProductResponse)
async def get_product(product_id: str, user: dict = Depends(get_current_user)):
    product = await db.products.find_one({"id": product_id}, {"_id": 0})
    if not product:
        raise HTTPException(status_code=404, detail="Produs negăsit")
    return ProductResponse(**product)

@api_router.put("/products/{product_id}", response_model=ProductResponse)
async def update_product(product_id: str, product: ProductUpdate, user: dict = Depends(require_admin)):
    update_data = {k: v for k, v in product.model_dump().items() if v is not None}
    if not update_data:
        raise HTTPException(status_code=400, detail="Nicio modificare")
    
    result = await db.products.update_one({"id": product_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Produs negăsit")
    
    updated = await db.products.find_one({"id": product_id}, {"_id": 0})
    return ProductResponse(**updated)

@api_router.delete("/products/{product_id}")
async def delete_product(product_id: str, user: dict = Depends(require_admin)):
    result = await db.products.delete_one({"id": product_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Produs negăsit")
    return {"message": "Produs șters"}

@api_router.get("/categories")
async def get_categories(user: dict = Depends(get_current_user)):
    categories = await db.products.distinct("categorie")
    return categories

# ==================== SUPPLIER ROUTES ====================

@api_router.post("/suppliers", response_model=SupplierResponse)
async def create_supplier(supplier: SupplierCreate, user: dict = Depends(require_admin)):
    supplier_doc = {
        "id": str(uuid.uuid4()),
        **supplier.model_dump(),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.suppliers.insert_one(supplier_doc)
    return SupplierResponse(**{k: v for k, v in supplier_doc.items() if k != "_id"})

@api_router.get("/suppliers", response_model=List[SupplierResponse])
async def get_suppliers(user: dict = Depends(get_current_user)):
    suppliers = await db.suppliers.find({}, {"_id": 0}).to_list(1000)
    return [SupplierResponse(**s) for s in suppliers]

@api_router.get("/suppliers/{supplier_id}", response_model=SupplierResponse)
async def get_supplier(supplier_id: str, user: dict = Depends(get_current_user)):
    supplier = await db.suppliers.find_one({"id": supplier_id}, {"_id": 0})
    if not supplier:
        raise HTTPException(status_code=404, detail="Furnizor negăsit")
    return SupplierResponse(**supplier)

@api_router.put("/suppliers/{supplier_id}", response_model=SupplierResponse)
async def update_supplier(supplier_id: str, supplier: SupplierCreate, user: dict = Depends(require_admin)):
    result = await db.suppliers.update_one({"id": supplier_id}, {"$set": supplier.model_dump()})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Furnizor negăsit")
    updated = await db.suppliers.find_one({"id": supplier_id}, {"_id": 0})
    return SupplierResponse(**updated)

@api_router.delete("/suppliers/{supplier_id}")
async def delete_supplier(supplier_id: str, user: dict = Depends(require_admin)):
    result = await db.suppliers.delete_one({"id": supplier_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Furnizor negăsit")
    return {"message": "Furnizor șters"}

# ==================== SALES ROUTES ====================

async def generate_bon_number():
    today = datetime.now(timezone.utc).strftime("%Y%m%d")
    count = await db.sales.count_documents({"numar_bon": {"$regex": f"^BON-{today}"}})
    return f"BON-{today}-{str(count + 1).zfill(4)}"

@api_router.post("/sales", response_model=SaleResponse)
async def create_sale(sale: SaleCreate, user: dict = Depends(get_current_user)):
    # P2: Duplicate prevention via transaction_id
    if sale.transaction_id:
        existing = await db.sales.find_one({"transaction_id": sale.transaction_id}, {"_id": 0})
        if existing:
            logger.warning(f"[SALE] Duplicate blocked: txn={sale.transaction_id}, user={user['username']}")
            return SaleResponse(**existing)
    
    # Generate bon number
    numar_bon = await generate_bon_number()
    
    # Get casier name
    casier = await db.users.find_one({"id": sale.casier_id}, {"_id": 0})
    casier_nume = casier["full_name"] if casier else "Necunoscut"
    
    sale_doc = {
        "id": str(uuid.uuid4()),
        "numar_bon": numar_bon,
        **sale.model_dump(),
        "transaction_id": sale.transaction_id or str(uuid.uuid4()),
        "casier_nume": casier_nume,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    # P0: Atomic stock deduction - only deduct after sale is validated
    stock_changes = []
    for item in sale.items:
        result = await db.products.update_one(
            {"id": item.product_id},
            {"$inc": {"stoc": -item.cantitate}}
        )
        stock_changes.append({
            "product_id": item.product_id,
            "nume": item.nume,
            "cantitate": -item.cantitate,
            "matched": result.matched_count
        })
    
    await db.sales.insert_one(sale_doc)
    
    # P3: Professional logging
    logger.info(
        f"[SALE] #{numar_bon} | Total: {sale.total} RON | "
        f"Plata: {sale.metoda_plata} | Fiscal: {sale.fiscal_status} | "
        f"Items: {len(sale.items)} | Casier: {casier_nume} | "
        f"TXN: {sale_doc['transaction_id']}"
    )
    for sc in stock_changes:
        if sc["matched"] == 0:
            logger.error(f"[STOCK] Product not found for deduction: {sc['product_id']} ({sc['nume']})")
    
    return SaleResponse(**{k: v for k, v in sale_doc.items() if k != "_id"})

@api_router.get("/sales", response_model=List[SaleResponse])
async def get_sales(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    user: dict = Depends(get_current_user)
):
    query = {}
    if start_date:
        query["created_at"] = {"$gte": start_date}
    if end_date:
        if "created_at" in query:
            query["created_at"]["$lte"] = end_date
        else:
            query["created_at"] = {"$lte": end_date}
    
    sales = await db.sales.find(query, {"_id": 0}).sort("created_at", -1).to_list(10000)
    return [SaleResponse(**s) for s in sales]

@api_router.get("/sales/{sale_id}", response_model=SaleResponse)
async def get_sale(sale_id: str, user: dict = Depends(get_current_user)):
    sale = await db.sales.find_one({"id": sale_id}, {"_id": 0})
    if not sale:
        raise HTTPException(status_code=404, detail="Vânzare negăsită")
    return SaleResponse(**sale)

@api_router.post("/sales/{sale_id}/cancel")
async def cancel_sale(sale_id: str, user: dict = Depends(get_current_user)):
    """Cancel a sale and restore stock"""
    sale = await db.sales.find_one({"id": sale_id}, {"_id": 0})
    if not sale:
        raise HTTPException(status_code=404, detail="Vânzare negăsită")
    
    if sale.get("fiscal_status") == "cancelled":
        return {"message": "Vânzarea este deja anulată"}
    
    # Restore stock for each item
    for item in sale.get("items", []):
        await db.products.update_one(
            {"id": item["product_id"]},
            {"$inc": {"stoc": item["cantitate"]}}
        )
    
    # Mark as cancelled
    await db.sales.update_one(
        {"id": sale_id},
        {"$set": {"fiscal_status": "cancelled"}}
    )
    
    logger.info(f"Sale {sale_id} cancelled, stock restored")
    return {"message": "Vânzare anulată, stoc restaurat"}

# ==================== SETTINGS ROUTES ====================

@api_router.get("/settings/fiscal")
async def get_fiscal_settings(user: dict = Depends(get_current_user)):
    """Get fiscal printer settings"""
    settings = await db.settings.find_one({"key": "fiscal"}, {"_id": 0})
    if not settings:
        return {
            "fiscal_mode": False,
            "bridge_url": "http://localhost:5555",
            "auto_print": True
        }
    return settings.get("value", {})

@api_router.post("/settings/fiscal")
async def update_fiscal_settings(data: dict, user: dict = Depends(require_admin)):
    """Update fiscal printer settings"""
    await db.settings.update_one(
        {"key": "fiscal"},
        {"$set": {"key": "fiscal", "value": data, "updated_at": datetime.now(timezone.utc).isoformat()}},
        upsert=True
    )
    return {"message": "Setări fiscale salvate", "settings": data}

# ==================== NIR ROUTES ====================

async def generate_nir_number():
    today = datetime.now(timezone.utc).strftime("%Y%m%d")
    count = await db.nirs.count_documents({"numar_nir": {"$regex": f"^NIR-{today}"}})
    return f"NIR-{today}-{str(count + 1).zfill(4)}"

@api_router.post("/nir", response_model=NIRResponse)
async def create_nir(nir: NIRCreate, user: dict = Depends(require_admin)):
    numar_nir = await generate_nir_number()
    
    # Get supplier name
    supplier = await db.suppliers.find_one({"id": nir.furnizor_id}, {"_id": 0})
    if not supplier:
        raise HTTPException(status_code=404, detail="Furnizor negăsit")
    
    nir_doc = {
        "id": str(uuid.uuid4()),
        "numar_nir": numar_nir,
        **nir.model_dump(),
        "furnizor_nume": supplier["nume"],
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    # Update stock for each item
    for item in nir.items:
        await db.products.update_one(
            {"id": item.product_id},
            {"$inc": {"stoc": item.cantitate}}
        )
    
    await db.nirs.insert_one(nir_doc)
    return NIRResponse(**{k: v for k, v in nir_doc.items() if k != "_id"})


@api_router.post("/nir/from-pdf")
async def create_nir_from_pdf(data: dict, user: dict = Depends(require_admin)):
    """Create NIR from PDF import - auto-creates new products for unmatched items"""
    furnizor_id = data.get("furnizor_id")
    numar_factura = data.get("numar_factura")
    items = data.get("items", [])

    if not furnizor_id or not numar_factura or not items:
        raise HTTPException(status_code=400, detail="Date incomplete")

    supplier = await db.suppliers.find_one({"id": furnizor_id}, {"_id": 0})
    if not supplier:
        raise HTTPException(status_code=404, detail="Furnizor negăsit")

    numar_nir = await generate_nir_number()
    nir_items = []
    created_products = []

    for item in items:
        product_id = item.get("product_id")
        denumire = item.get("denumire", item.get("nume", ""))
        cantitate = float(item.get("cantitate", 0))
        pret_achizitie = float(item.get("pret_achizitie", 0))
        um = item.get("um", "buc")

        if not product_id:
            # Auto-create new product
            new_product = {
                "id": str(uuid.uuid4()),
                "nume": denumire,
                "categorie": "Necategorisit",
                "cod_bare": "",
                "pret_achizitie": pret_achizitie,
                "pret_vanzare": round(pret_achizitie * 1.3, 2),  # 30% markup default
                "tva": 19,
                "unitate": um if um in ["buc", "sac", "kg", "metru", "litru", "rola"] else "buc",
                "stoc": cantitate,
                "stoc_minim": 5,
                "descriere": "",
                "furnizor_id": furnizor_id,
                "created_at": datetime.now(timezone.utc).isoformat()
            }
            await db.products.insert_one(new_product)
            product_id = new_product["id"]
            created_products.append({"product_id": product_id, "nume": denumire, "cod_bare": ""})
        else:
            # Update existing product stock
            await db.products.update_one(
                {"id": product_id},
                {"$inc": {"stoc": cantitate}}
            )
            prod = await db.products.find_one({"id": product_id}, {"_id": 0, "nume": 1, "cod_bare": 1})
            created_products.append({"product_id": product_id, "nume": prod["nume"] if prod else denumire, "cod_bare": prod.get("cod_bare", "") if prod else ""})

        nir_items.append({
            "product_id": product_id,
            "nume": denumire,
            "cantitate": cantitate,
            "pret_achizitie": pret_achizitie
        })

    total = sum(i["cantitate"] * i["pret_achizitie"] for i in nir_items)
    nir_doc = {
        "id": str(uuid.uuid4()),
        "numar_nir": numar_nir,
        "furnizor_id": furnizor_id,
        "numar_factura": numar_factura,
        "items": nir_items,
        "total": total,
        "furnizor_nume": supplier["nume"],
        "created_at": datetime.now(timezone.utc).isoformat()
    }

    await db.nirs.insert_one(nir_doc)
    nir_response = {k: v for k, v in nir_doc.items() if k != "_id"}
    return {
        "nir": nir_response,
        "created_products": created_products,
        "products_created_count": len([p for p in items if not p.get("product_id")]),
        "products_updated_count": len([p for p in items if p.get("product_id")])
    }

@api_router.get("/nir", response_model=List[NIRResponse])
async def get_nirs(user: dict = Depends(get_current_user)):
    nirs = await db.nirs.find({}, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return [NIRResponse(**n) for n in nirs]


@api_router.post("/products/bulk-barcode")
async def bulk_update_barcodes(data: dict, user: dict = Depends(require_admin)):
    """Update barcodes for multiple products at once (after NIR)"""
    updates = data.get("updates", [])
    if not updates:
        raise HTTPException(status_code=400, detail="Nu sunt actualizări")

    updated = 0
    for upd in updates:
        product_id = upd.get("product_id")
        barcode = upd.get("cod_bare", "").strip()
        if product_id and barcode:
            result = await db.products.update_one(
                {"id": product_id},
                {"$set": {"cod_bare": barcode}}
            )
            if result.modified_count > 0:
                updated += 1

    return {"updated": updated, "total": len(updates)}


@api_router.get("/nir/test-invoices")
async def list_test_invoices(user: dict = Depends(require_admin)):
    """List available test invoice PDFs"""
    invoice_dir = os.path.join(os.path.dirname(__file__), "static", "test_invoices")
    if not os.path.exists(invoice_dir):
        return {"invoices": []}
    files = [f for f in os.listdir(invoice_dir) if f.endswith('.pdf')]
    return {"invoices": files}


@api_router.get("/nir/test-invoices/{filename}")
async def download_test_invoice(filename: str, user: dict = Depends(require_admin)):
    """Download a test invoice PDF"""
    invoice_dir = os.path.join(os.path.dirname(__file__), "static", "test_invoices")
    filepath = os.path.join(invoice_dir, filename)
    if not os.path.exists(filepath) or not filename.endswith('.pdf'):
        raise HTTPException(status_code=404, detail="Fișier negăsit")
    return FileResponse(filepath, media_type="application/pdf", filename=filename)


@api_router.post("/nir/parse-pdf")
async def parse_nir_pdf(file: UploadFile = File(...), user: dict = Depends(require_admin)):
    """Parse a supplier invoice PDF and extract items for NIR creation"""
    if not file.filename.lower().endswith('.pdf'):
        raise HTTPException(status_code=400, detail="Fișierul trebuie să fie PDF")

    content = await file.read()
    if len(content) > 10 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Fișierul este prea mare (max 10MB)")

    try:
        doc = fitz.open(stream=content, filetype="pdf")
    except Exception:
        raise HTTPException(status_code=400, detail="Fișierul PDF nu a putut fi deschis")

    all_text = ""
    all_blocks = []
    for page in doc:
        all_text += page.get_text("text") + "\n"
        for block in page.get_text("blocks"):
            if block[6] == 0:  # text blocks only
                all_blocks.append(block[4])
    doc.close()

    # Try to extract invoice number
    invoice_number = ""
    inv_patterns = [
        r'(?:factura|fact\.?|fv|invoice)\s*(?:nr\.?|no\.?|numar|#)?\s*[:.]?\s*\n?\s*([A-Z0-9\-\/\s]+)',
        r'(?:TOP|SER|FV|FC)\s+(\d{5,})',
        r'(?:seria?\s+[A-Z]+\s+nr\.?\s*)(\d+)',
    ]
    for pattern in inv_patterns:
        match = re.search(pattern, all_text, re.IGNORECASE)
        if match:
            invoice_number = match.group(1).strip().split('\n')[0].strip()
            break

    # Try to extract supplier name
    supplier_name = ""
    # Strategy 1: Find company names (S.R.L., S.A.) that are NOT "ANDREPAU"
    company_matches = re.findall(r'([\w\s\.\-]+\bS\.?R\.?L\.?\b|[\w\s\.\-]+\bS\.?A\.?\b|[\w\s\.\-]+\bL\.?T\.?D\.?\b)', all_text)
    for name in company_matches:
        clean = name.strip()
        if clean and 'ANDREPAU' not in clean.upper() and len(clean) > 5:
            # Remove leading numbers, dates, etc.
            clean = re.sub(r'^\d+[\s.]*', '', clean).strip()
            if clean and len(clean) > 5:
                supplier_name = clean
                break
    # Strategy 2: After "Furnizor:" scan lines for a company-like name
    if not supplier_name:
        lines = all_text.split('\n')
        in_supplier_section = False
        for line in lines:
            stripped = line.strip()
            if re.match(r'(?i)furnizor', stripped):
                in_supplier_section = True
                continue
            if in_supplier_section and stripped:
                if 'cumparator' in stripped.lower() or 'andrepau' in stripped.lower():
                    continue
                if len(stripped) > 3 and not stripped.startswith('CUI') and not stripped.startswith('Reg'):
                    supplier_name = stripped
                    break

    # Extract items using blocks (primary strategy for e-Factura PDFs)
    extracted_items = _extract_items_from_blocks(all_blocks)

    # Fallback to line-by-line parsing if blocks didn't work
    if not extracted_items:
        extracted_items = _extract_items_from_lines(all_text)

    # Match with existing products
    products = await db.products.find({}, {"_id": 0, "id": 1, "nume": 1, "cod_bare": 1, "pret_achizitie": 1}).to_list(100000)
    matched_items = []

    for item in extracted_items:
        best_match = _find_best_product_match(item["denumire"], products)
        matched_items.append({
            "denumire_pdf": item["denumire"],
            "cantitate": item["cantitate"],
            "pret_unitar": item["pret_unitar"],
            "um": item.get("um", "buc"),
            "valoare": item.get("valoare", round(item["cantitate"] * item["pret_unitar"], 2)),
            "product_id": best_match["id"] if best_match else None,
            "product_nume": best_match["nume"] if best_match else None,
            "match_confidence": best_match["confidence"] if best_match else 0,
        })

    return {
        "invoice_number": invoice_number,
        "supplier_name": supplier_name,
        "items": matched_items,
        "raw_text_preview": all_text[:2000],
        "total_items": len(matched_items)
    }


def _extract_items_from_blocks(blocks: list) -> list:
    """Extract items from PyMuPDF text blocks - works with e-Factura and standard Romanian invoices.
    Each product row is typically a single block with fields separated by newlines:
    'Nr\\nDenumire\\nUM\\nCantitate\\nPretUnitar\\nValoare\\nTVA\\n'
    """
    items = []
    um_set = {"buc", "sac", "kg", "m", "ml", "m2", "mp", "mc", "litru", "l", "rola", "set", "to", "pach", "pac", "fl", "cutie", "cmp"}

    for block_text in blocks:
        parts = [p.strip() for p in block_text.strip().split('\n') if p.strip()]
        if len(parts) < 4:
            continue

        # Check if first part is a row number (1, 2, 3, etc.)
        first = parts[0]
        if not re.match(r'^\d{1,4}$', first):
            continue

        row_num = int(first)
        if row_num < 1 or row_num > 999:
            continue

        # Try to parse: Nr, Denumire, UM, Cantitate, PretUnitar, [Valoare], [TVA]
        # Find the UM field - it tells us where quantity starts
        um_idx = None
        for i, part in enumerate(parts[1:], 1):
            if part.lower() in um_set:
                um_idx = i
                break

        if um_idx is not None and um_idx + 2 < len(parts):
            # Standard format: Nr | Denumire | UM | Cant | Pret | Valoare | TVA
            name = ' '.join(parts[1:um_idx])
            um = parts[um_idx].lower()
            qty = _parse_number(parts[um_idx + 1])
            price = _parse_number(parts[um_idx + 2])
            valoare = _parse_number(parts[um_idx + 3]) if um_idx + 3 < len(parts) else round(qty * price, 2)

            if qty > 0 and price > 0 and len(name) > 2:
                items.append({
                    "denumire": name,
                    "um": um if um in um_set else "buc",
                    "cantitate": qty,
                    "pret_unitar": price,
                    "valoare": valoare
                })
        else:
            # Try without UM: Nr | Denumire | Cant | Pret | Valoare
            # Find where numbers start (after name)
            num_start = None
            for i in range(1, len(parts)):
                cleaned = parts[i].replace(',', '.').replace(' ', '')
                try:
                    float(cleaned)
                    num_start = i
                    break
                except ValueError:
                    continue

            if num_start is not None and num_start > 1:
                name = ' '.join(parts[1:num_start])
                numbers = []
                for p in parts[num_start:]:
                    val = _parse_number(p)
                    if val > 0:
                        numbers.append(val)

                if len(numbers) >= 2:
                    qty = numbers[0]
                    price = numbers[1]
                    valoare = numbers[2] if len(numbers) > 2 else round(qty * price, 2)
                    if len(name) > 2:
                        items.append({
                            "denumire": name,
                            "um": "buc",
                            "cantitate": qty,
                            "pret_unitar": price,
                            "valoare": valoare
                        })

    return items


def _extract_items_from_lines(text: str) -> list:
    """Fallback: extract items from text line by line for non-block PDFs"""
    items = []
    lines = text.split('\n')
    lines = [ln.strip() for ln in lines if ln.strip()]

    skip_words = {'denumire', 'produs', 'nr.crt', 'nr. crt', 'cantitate', 'total general',
                  'subtotal', 'tva', 'baza impozabila', 'de plata', 'cont', 'banca',
                  'factura', 'seria', 'furnizor', 'cumparator', 'adresa', 'delegat',
                  'cota tva', 'valoare tva', 'index', 'livrare', 'intocmit', 'semnatura'}

    seen_names = set()
    i = 0
    while i < len(lines):
        line = lines[i]
        lower = line.lower()
        if any(h in lower for h in skip_words):
            i += 1
            continue

        # Check if line starts with a row number
        m = re.match(r'^(\d{1,3})\s*$', line)
        if m:
            row_num = int(m.group(1))
            if 1 <= row_num <= 999 and i + 3 < len(lines):
                # Next lines should be: name, [um], qty, price, value
                name = lines[i + 1].strip()
                if name.lower() in skip_words or re.match(r'^\d+[.,]?\d*$', name):
                    i += 1
                    continue

                # Look ahead for numbers
                remaining = lines[i + 2:i + 8]
                um = "buc"
                numbers = []
                for r_line in remaining:
                    r_clean = r_line.strip().lower()
                    if r_clean in {"buc", "sac", "kg", "m", "ml", "m2", "mp", "mc", "litru", "l", "rola", "set"}:
                        um = r_clean
                        continue
                    val = _parse_number(r_line)
                    if val > 0:
                        numbers.append(val)
                    elif r_line.strip() and not re.match(r'^[\d.,\s]+$', r_line.strip()):
                        break

                if len(numbers) >= 2 and name not in seen_names:
                    seen_names.add(name)
                    items.append({
                        "denumire": name,
                        "um": um,
                        "cantitate": numbers[0],
                        "pret_unitar": numbers[1],
                        "valoare": numbers[2] if len(numbers) > 2 else round(numbers[0] * numbers[1], 2)
                    })

        i += 1

    return items


def _parse_number(s: str) -> float:
    """Parse a number string that might use comma as decimal separator"""
    s = s.strip()
    # If has both . and , => the last one is the decimal separator
    if '.' in s and ',' in s:
        if s.rindex('.') > s.rindex(','):
            s = s.replace(',', '')
        else:
            s = s.replace('.', '').replace(',', '.')
    elif ',' in s:
        s = s.replace(',', '.')
    try:
        return float(s)
    except ValueError:
        return 0.0


def _find_best_product_match(pdf_name: str, products: list) -> dict | None:
    """Find best matching product by name similarity"""
    if not pdf_name or not products:
        return None

    pdf_lower = pdf_name.lower().strip()
    best = None
    best_score = 0

    for p in products:
        prod_lower = p["nume"].lower().strip()

        # Exact match
        if pdf_lower == prod_lower:
            return {"id": p["id"], "nume": p["nume"], "confidence": 100}

        # Containment check
        if pdf_lower in prod_lower or prod_lower in pdf_lower:
            score = 80
            if score > best_score:
                best_score = score
                best = {"id": p["id"], "nume": p["nume"], "confidence": score}
            continue

        # Word overlap
        pdf_words = set(pdf_lower.split())
        prod_words = set(prod_lower.split())
        if pdf_words and prod_words:
            overlap = len(pdf_words & prod_words)
            total = max(len(pdf_words), len(prod_words))
            score = int((overlap / total) * 70)
            if score > best_score and score >= 30:
                best_score = score
                best = {"id": p["id"], "nume": p["nume"], "confidence": score}

    return best


# ==================== STOCK ROUTES ====================

@api_router.get("/stock/dashboard")
async def get_stock_dashboard(user: dict = Depends(get_current_user)):
    total_products = await db.products.count_documents({})
    
    # Products with low stock
    pipeline_low = [
        {"$match": {"$expr": {"$lte": ["$stoc", "$stoc_minim"]}, "stoc": {"$gt": 0}}},
        {"$count": "count"}
    ]
    low_stock_result = await db.products.aggregate(pipeline_low).to_list(1)
    low_stock = low_stock_result[0]["count"] if low_stock_result else 0
    
    # Products out of stock
    out_of_stock = await db.products.count_documents({"stoc": {"$lte": 0}})
    
    # Total stock value
    pipeline_value = [
        {"$group": {"_id": None, "total": {"$sum": {"$multiply": ["$stoc", "$pret_achizitie"]}}}}
    ]
    value_result = await db.products.aggregate(pipeline_value).to_list(1)
    total_value = value_result[0]["total"] if value_result else 0
    
    return {
        "total_products": total_products,
        "low_stock": low_stock,
        "out_of_stock": out_of_stock,
        "total_value": round(total_value, 2)
    }

@api_router.get("/stock/alerts")
async def get_stock_alerts(user: dict = Depends(get_current_user)):
    pipeline = [
        {"$match": {"$expr": {"$lte": ["$stoc", "$stoc_minim"]}}},
        {"$addFields": {
            "severity": {
                "$cond": [{"$lte": ["$stoc", 0]}, "critical", "warning"]
            },
            "deficit": {"$subtract": ["$stoc_minim", "$stoc"]}
        }},
        {"$sort": {"severity": 1, "deficit": -1}},
        {"$project": {"_id": 0}}
    ]
    alerts = await db.products.aggregate(pipeline).to_list(200)
    return alerts

# ==================== REPORTS ROUTES ====================

@api_router.get("/reports/sales")
async def get_sales_report(
    period: str = "today",  # today, week, month, year
    user: dict = Depends(get_current_user)
):
    now = datetime.now(timezone.utc)
    
    if period == "today":
        start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    elif period == "week":
        start = now - timedelta(days=7)
    elif period == "month":
        start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    elif period == "year":
        start = now.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
    else:
        start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    
    pipeline = [
        {"$match": {"created_at": {"$gte": start.isoformat()}}},
        {"$group": {
            "_id": None,
            "total_sales": {"$sum": "$total"},
            "total_tva": {"$sum": "$tva_total"},
            "count": {"$sum": 1},
            "cash": {"$sum": "$suma_numerar"},
            "card": {"$sum": "$suma_card"}
        }}
    ]
    
    result = await db.sales.aggregate(pipeline).to_list(1)
    
    if result:
        return {
            "total_sales": round(result[0]["total_sales"], 2),
            "total_tva": round(result[0]["total_tva"], 2),
            "count": result[0]["count"],
            "cash": round(result[0]["cash"], 2),
            "card": round(result[0]["card"], 2)
        }
    return {"total_sales": 0, "total_tva": 0, "count": 0, "cash": 0, "card": 0}

@api_router.get("/reports/top-products")
async def get_top_products(limit: int = 10, user: dict = Depends(get_current_user)):
    pipeline = [
        {"$unwind": "$items"},
        {"$group": {
            "_id": "$items.product_id",
            "nume": {"$first": "$items.nume"},
            "total_cantitate": {"$sum": "$items.cantitate"},
            "total_valoare": {"$sum": {"$multiply": ["$items.cantitate", "$items.pret_unitar"]}}
        }},
        {"$sort": {"total_valoare": -1}},
        {"$limit": limit}
    ]
    
    results = await db.sales.aggregate(pipeline).to_list(limit)
    return [
        {
            "product_id": r["_id"],
            "nume": r["nume"],
            "total_cantitate": round(r["total_cantitate"], 2),
            "total_valoare": round(r["total_valoare"], 2)
        }
        for r in results
    ]

@api_router.get("/reports/top-categories")
async def get_top_categories(user: dict = Depends(get_current_user)):
    pipeline = [
        {"$unwind": "$items"},
        {"$lookup": {
            "from": "products",
            "localField": "items.product_id",
            "foreignField": "id",
            "as": "product"
        }},
        {"$unwind": {"path": "$product", "preserveNullAndEmptyArrays": True}},
        {"$group": {
            "_id": "$product.categorie",
            "total_valoare": {"$sum": {"$multiply": ["$items.cantitate", "$items.pret_unitar"]}}
        }},
        {"$sort": {"total_valoare": -1}},
        {"$limit": 10}
    ]
    
    results = await db.sales.aggregate(pipeline).to_list(10)
    return [
        {
            "categorie": r["_id"] or "Necunoscut",
            "total_valoare": round(r["total_valoare"], 2)
        }
        for r in results
    ]

@api_router.get("/reports/profit")
async def get_profit_report(period: str = "month", user: dict = Depends(require_admin)):
    now = datetime.now(timezone.utc)
    
    if period == "today":
        start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    elif period == "week":
        start = now - timedelta(days=7)
    elif period == "month":
        start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    else:
        start = now.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
    
    # Get all sales in period
    sales = await db.sales.find(
        {"created_at": {"$gte": start.isoformat()}},
        {"_id": 0}
    ).to_list(10000)
    
    total_vanzari = 0
    total_cost = 0
    
    for sale in sales:
        for item in sale["items"]:
            total_vanzari += item["cantitate"] * item["pret_unitar"]
            # Get product cost
            product = await db.products.find_one({"id": item["product_id"]}, {"_id": 0})
            if product:
                total_cost += item["cantitate"] * product["pret_achizitie"]
    
    profit = total_vanzari - total_cost
    margin = (profit / total_vanzari * 100) if total_vanzari > 0 else 0
    
    return {
        "total_vanzari": round(total_vanzari, 2),
        "total_cost": round(total_cost, 2),
        "profit": round(profit, 2),
        "margin_percent": round(margin, 2)
    }

@api_router.get("/reports/daily-sales")
async def get_daily_sales(days: int = 30, user: dict = Depends(get_current_user)):
    start = datetime.now(timezone.utc) - timedelta(days=days)
    
    pipeline = [
        {"$match": {"created_at": {"$gte": start.isoformat()}}},
        {"$addFields": {
            "date": {"$substr": ["$created_at", 0, 10]}
        }},
        {"$group": {
            "_id": "$date",
            "total": {"$sum": "$total"},
            "count": {"$sum": 1}
        }},
        {"$sort": {"_id": 1}}
    ]
    
    results = await db.sales.aggregate(pipeline).to_list(days)
    return [{"date": r["_id"], "total": round(r["total"], 2), "count": r["count"]} for r in results]

# ==================== CASH OPERATIONS ROUTES ====================

@api_router.post("/cash-operations")
async def create_cash_operation(data: dict, user: dict = Depends(get_current_user)):
    """Save a cash operation (Cash In, Cash Out, Report X, Report Z)"""
    operation = {
        "id": str(uuid.uuid4()),
        "type": data.get("type"),  # CASH_IN, CASH_OUT, REPORT_X, REPORT_Z
        "amount": data.get("amount", 0),
        "description": data.get("description", ""),
        "operator_id": data.get("operator_id"),
        "operator_name": data.get("operator_name"),
        "timestamp": datetime.now(timezone.utc),
        "date_str": datetime.now(timezone.utc).strftime("%Y-%m-%d")
    }
    
    await db.cash_operations.insert_one(operation)
    operation.pop("_id", None)
    logger.info(f"[CASH] {operation['type']} | Amount: {operation['amount']} RON | By: {operation['operator_name']} | Desc: {operation['description']}")
    return operation

@api_router.get("/cash-operations/history")
async def get_cash_operations_history(
    limit: int = 50,
    date: Optional[str] = None,
    user: dict = Depends(get_current_user)
):
    """Get cash operations history"""
    query = {}
    if date:
        query["date_str"] = date
    
    operations = await db.cash_operations.find(query, {"_id": 0}).sort("timestamp", -1).to_list(limit)
    return {"operations": operations}

@api_router.get("/cash-operations/daily-stats")
async def get_daily_cash_stats(user: dict = Depends(get_current_user)):
    """Get daily statistics for cash operations"""
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    today_start = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0)
    
    # Sales totals by payment method
    sales_pipeline = [
        {"$match": {"created_at": {"$gte": today_start}}},
        {"$group": {
            "_id": "$metoda_plata",
            "total": {"$sum": "$total"},
            "count": {"$sum": 1}
        }}
    ]
    sales_by_method = await db.sales.aggregate(sales_pipeline).to_list(10)
    
    total_cash = 0
    total_card = 0
    total_voucher = 0
    receipts_count = 0
    
    for s in sales_by_method:
        receipts_count += s.get("count", 0)
        if s["_id"] in ["numerar", "cash"]:
            total_cash = s.get("total", 0)
        elif s["_id"] in ["card"]:
            total_card = s.get("total", 0)
        elif s["_id"] in ["tichete", "voucher"]:
            total_voucher = s.get("total", 0)
    
    # Cash In/Out totals
    cash_ops_pipeline = [
        {"$match": {"date_str": today}},
        {"$group": {
            "_id": "$type",
            "total": {"$sum": "$amount"}
        }}
    ]
    cash_ops = await db.cash_operations.aggregate(cash_ops_pipeline).to_list(10)
    
    cash_in = 0
    cash_out = 0
    for op in cash_ops:
        if op["_id"] == "CASH_IN":
            cash_in = op.get("total", 0)
        elif op["_id"] == "CASH_OUT":
            cash_out = op.get("total", 0)
    
    return {
        "totalCash": round(total_cash, 2),
        "totalCard": round(total_card, 2),
        "totalVoucher": round(total_voucher, 2),
        "cashIn": round(cash_in, 2),
        "cashOut": round(cash_out, 2),
        "receiptsCount": receipts_count
    }

@api_router.get("/daily/opening-summary")
async def get_opening_summary(user: dict = Depends(get_current_user)):
    """Dashboard deschidere zi - summary complet"""
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    today_start = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0)
    
    # Cash balance from today's operations
    cash_ops = await db.cash_operations.aggregate([
        {"$match": {"date_str": today}},
        {"$group": {"_id": "$type", "total": {"$sum": "$amount"}}}
    ]).to_list(10)
    cash_in = sum(op["total"] for op in cash_ops if op["_id"] == "CASH_IN")
    cash_out = sum(op["total"] for op in cash_ops if op["_id"] == "CASH_OUT")
    
    # Today's cash sales
    cash_sales = await db.sales.aggregate([
        {"$match": {"created_at": {"$gte": today_start.isoformat()}, "metoda_plata": {"$in": ["numerar", "cash"]}}},
        {"$group": {"_id": None, "total": {"$sum": "$total"}, "count": {"$sum": 1}}}
    ]).to_list(1)
    cash_from_sales = cash_sales[0]["total"] if cash_sales else 0
    sales_count = cash_sales[0]["count"] if cash_sales else 0
    
    # Total sales today (all methods)
    all_sales = await db.sales.aggregate([
        {"$match": {"created_at": {"$gte": today_start.isoformat()}}},
        {"$group": {"_id": None, "total": {"$sum": "$total"}, "count": {"$sum": 1}}}
    ]).to_list(1)
    total_sales = all_sales[0]["total"] if all_sales else 0
    total_sales_count = all_sales[0]["count"] if all_sales else 0
    
    # Cash register balance estimate
    sold_casa = cash_in - cash_out + cash_from_sales
    
    # Bridge status
    last_poll = await db.fiscal_bridge_status.find_one({"key": "last_poll"}, {"_id": 0})
    bridge_connected = False
    if last_poll:
        last_time = datetime.fromisoformat(last_poll["timestamp"])
        bridge_connected = (datetime.now(timezone.utc) - last_time).total_seconds() < 30
    
    # Held orders
    await expire_old_held_orders()
    held_count = await db.held_orders.count_documents({"status": "active"})
    
    # Stock alerts count
    alert_count = await db.products.count_documents({"$expr": {"$lte": ["$stoc", "$stoc_minim"]}})
    out_of_stock = await db.products.count_documents({"stoc": {"$lte": 0}})
    
    return {
        "data": today,
        "sold_casa": round(sold_casa, 2),
        "cash_in": round(cash_in, 2),
        "cash_out": round(cash_out, 2),
        "vanzari_numerar": round(cash_from_sales, 2),
        "total_vanzari": round(total_sales, 2),
        "numar_vanzari": total_sales_count,
        "bridge_connected": bridge_connected,
        "comenzi_hold": held_count,
        "alerte_stoc": alert_count,
        "fara_stoc": out_of_stock,
        "ora": datetime.now(timezone.utc).strftime("%H:%M")
    }


# ==================== TVA MANAGEMENT ====================

@api_router.get("/settings/tva")
async def get_tva_settings():
    """Returneaza cotele TVA configurate"""
    return {
        "cote_tva": [
            {"cod": "A", "procent": 21.0, "descriere": "Standard (21%)"},
            {"cod": "B", "procent": 11.0, "descriere": "Redusa (11%)"},
            {"cod": "C", "procent": 0.0, "descriere": "Scutit/Export (0%)"},
            {"cod": "D", "procent": 9.0, "descriere": "Tranzitorie locuinte (9%)"},
        ],
        "nota_legala": "Conform Legea 141/2025, valabil de la 1 august 2025"
    }

@api_router.post("/products/update-tva")
async def update_products_tva(
    data: dict,
    user: dict = Depends(get_current_user)
):
    """Actualizeaza TVA-ul pe toate produsele (admin only)"""
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Doar administratorul poate modifica TVA")
    
    old_tva = data.get("old_tva", 19.0)
    new_tva = data.get("new_tva", 21.0)
    
    result = await db.products.update_many(
        {"tva": old_tva},
        {"$set": {"tva": new_tva}}
    )
    
    return {
        "message": f"TVA actualizat de la {old_tva}% la {new_tva}%",
        "products_updated": result.modified_count
    }

# ==================== BRIDGE DOWNLOAD ====================

@api_router.get("/bridge/download")
async def download_bridge_zip(user: dict = Depends(get_current_user)):
    """Download Bridge Service as ZIP (fiscal_bridge.py + install + start scripts)"""
    return _create_bridge_zip()

@api_router.get("/bridge/download-direct")
async def download_bridge_direct(token: str = None):
    """Download Bridge Service with token in URL - for direct browser download"""
    if not token:
        raise HTTPException(status_code=401, detail="Token necesar")
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
    except:
        raise HTTPException(status_code=401, detail="Token invalid")
    return _create_bridge_zip()

def _create_bridge_zip():
    import zipfile
    bridge_dir = Path(__file__).parent
    files_to_zip = {
        "fiscal_bridge.py": bridge_dir / "fiscal_bridge.py",
        "install_bridge.bat": bridge_dir / "install_bridge.bat",
        "start_bridge.bat": bridge_dir / "start_bridge.bat",
        "ACTUALIZEAZA_BRIDGE.bat": bridge_dir / "actualizeaza_bridge.bat",
    }
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
        for name, path in files_to_zip.items():
            if path.exists():
                zf.write(path, name)
    zip_buffer.seek(0)
    return StreamingResponse(
        zip_buffer,
        media_type="application/zip",
        headers={"Content-Disposition": "attachment; filename=ANDREPAU_Bridge_Service.zip"}
    )

# ==================== FISCAL JOB QUEUE (Bridge Cloud Relay) ====================
# PWA -> Backend (queue job) -> Bridge polls -> Executes -> Reports result -> PWA gets result

@api_router.post("/fiscal/queue")
async def queue_fiscal_job(data: dict, user: dict = Depends(get_current_user)):
    """PWA trimite comanda fiscala in coada - bridge-ul o va prelua"""
    job_id = str(uuid.uuid4())
    job = {
        "job_id": job_id,
        "type": data.get("type", "receipt"),
        "data": data.get("data", {}),
        "status": "pending",
        "result": None,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "created_by": user.get("username", "unknown"),
        "completed_at": None
    }
    await db.fiscal_jobs.insert_one(job)
    logger.info(f"[FISCAL] Job queued: {job['type']} | ID: {job_id} | By: {user.get('username')}")
    return {"job_id": job_id, "status": "pending"}

@api_router.get("/fiscal/pending")
async def get_pending_jobs(bridge_key: str = None):
    """Bridge-ul poll-uieste pentru joburi noi (nu necesita auth JWT)"""
    # Simple key auth for bridge
    jobs = await db.fiscal_jobs.find(
        {"status": "pending"},
        {"_id": 0}
    ).sort("created_at", 1).limit(1).to_list(1)
    
    if jobs:
        # Mark as processing
        job = jobs[0]
        await db.fiscal_jobs.update_one(
            {"job_id": job["job_id"], "status": "pending"},
            {"$set": {"status": "processing"}}
        )
        return {"job": job}
    return {"job": None}

@api_router.post("/fiscal/result/{job_id}")
async def report_fiscal_result(job_id: str, data: dict):
    """Bridge-ul raporteaza rezultatul executiei"""
    status = "completed" if data.get("success") else "failed"
    result = await db.fiscal_jobs.update_one(
        {"job_id": job_id},
        {"$set": {
            "status": status,
            "result": data,
            "completed_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Job negasit")
    logger.info(f"[FISCAL] Job result: {job_id} | Status: {status} | Error: {data.get('error', 'none')}")
    return {"ok": True}

@api_router.get("/fiscal/status/{job_id}")
async def get_fiscal_job_status(job_id: str):
    """PWA verifica statusul jobului fiscal"""
    job = await db.fiscal_jobs.find_one(
        {"job_id": job_id},
        {"_id": 0}
    )
    if not job:
        raise HTTPException(status_code=404, detail="Job negasit")
    return job

@api_router.get("/fiscal/bridge-status")
async def get_bridge_status():
    """Verifica daca bridge-ul e activ (a poll-uit recent)"""
    # Check if bridge polled in last 30 seconds
    last_poll = await db.fiscal_bridge_status.find_one(
        {"key": "last_poll"},
        {"_id": 0}
    )
    if last_poll:
        last_time = datetime.fromisoformat(last_poll["timestamp"])
        now = datetime.now(timezone.utc)
        connected = (now - last_time).total_seconds() < 30
        return {"connected": connected, "last_poll": last_poll["timestamp"]}
    return {"connected": False, "last_poll": None}

@api_router.post("/fiscal/bridge-ping")
async def bridge_ping():
    """Bridge-ul trimite ping ca sa anunte ca e activ"""
    await db.fiscal_bridge_status.update_one(
        {"key": "last_poll"},
        {"$set": {"key": "last_poll", "timestamp": datetime.now(timezone.utc).isoformat()}},
        upsert=True
    )
    return {"ok": True}


# ==================== HELD ORDERS (Stock Reservation) ====================

async def expire_old_held_orders():
    """Expire held orders older than 24 hours - stock STAYS deducted (not restored)"""
    cutoff = (datetime.now(timezone.utc) - timedelta(hours=24)).isoformat()
    expired_orders = await db.held_orders.find(
        {"status": "active", "created_at": {"$lt": cutoff}},
        {"_id": 0}
    ).to_list(100)
    
    for order in expired_orders:
        # Stock remains deducted - products are considered gone/sold
        await db.held_orders.update_one(
            {"id": order["id"]},
            {"$set": {"status": "expired", "expired_at": datetime.now(timezone.utc).isoformat()}}
        )
    
    if expired_orders:
        logger.info(f"Expired {len(expired_orders)} held orders, stock remains deducted")

@api_router.post("/held-orders")
async def create_held_order(data: dict, user: dict = Depends(get_current_user)):
    """Create a held order and reserve stock (deduct immediately)"""
    items = data.get("items", [])
    if not items:
        raise HTTPException(status_code=400, detail="Nu sunt produse in cos")
    
    held_order = {
        "id": str(uuid.uuid4()),
        "items": items,
        "discount": data.get("discount", 0),
        "created_by": user["id"],
        "created_by_name": user.get("full_name", user["username"]),
        "status": "active",
        "created_at": datetime.now(timezone.utc).isoformat(),
        "expires_at": (datetime.now(timezone.utc) + timedelta(hours=24)).isoformat()
    }
    
    # Deduct stock for each item (reserve)
    for item in items:
        await db.products.update_one(
            {"id": item["product_id"]},
            {"$inc": {"stoc": -item["cantitate"]}}
        )
    
    await db.held_orders.insert_one(held_order)
    held_order.pop("_id", None)
    logger.info(f"Held order {held_order['id']} created by {user['username']}, {len(items)} items reserved")
    return held_order

@api_router.get("/held-orders")
async def get_held_orders(user: dict = Depends(get_current_user)):
    """Get all active held orders (expires old ones first)"""
    await expire_old_held_orders()
    
    orders = await db.held_orders.find(
        {"status": "active"},
        {"_id": 0}
    ).sort("created_at", -1).to_list(100)
    return {"orders": orders, "total": len(orders)}

@api_router.post("/held-orders/{order_id}/restore")
async def restore_held_order(order_id: str, user: dict = Depends(get_current_user)):
    """Restore a held order to cart - restores stock (will be deducted again on sale)"""
    order = await db.held_orders.find_one({"id": order_id, "status": "active"}, {"_id": 0})
    if not order:
        raise HTTPException(status_code=404, detail="Comanda nu a fost gasita sau a expirat")
    
    # Restore stock (will be deducted again when sale is finalized)
    for item in order["items"]:
        await db.products.update_one(
            {"id": item["product_id"]},
            {"$inc": {"stoc": item["cantitate"]}}
        )
    
    await db.held_orders.update_one(
        {"id": order_id},
        {"$set": {"status": "restored", "restored_at": datetime.now(timezone.utc).isoformat()}}
    )
    
    logger.info(f"Held order {order_id} restored by {user['username']}")
    return order

@api_router.post("/held-orders/{order_id}/cancel")
async def cancel_held_order(order_id: str, user: dict = Depends(get_current_user)):
    """Cancel a held order and restore stock"""
    order = await db.held_orders.find_one({"id": order_id, "status": "active"}, {"_id": 0})
    if not order:
        raise HTTPException(status_code=404, detail="Comanda nu a fost gasita")
    
    for item in order["items"]:
        await db.products.update_one(
            {"id": item["product_id"]},
            {"$inc": {"stoc": item["cantitate"]}}
        )
    
    await db.held_orders.update_one(
        {"id": order_id},
        {"$set": {"status": "cancelled", "cancelled_at": datetime.now(timezone.utc).isoformat()}}
    )
    
    logger.info(f"Held order {order_id} cancelled by {user['username']}")
    return {"message": "Comanda anulata, stoc restaurat"}



# ==================== BACKUP ROUTE ====================

@api_router.get("/backup")
async def create_backup(user: dict = Depends(require_admin)):
    backup_data = {
        "created_at": datetime.now(timezone.utc).isoformat(),
        "products": await db.products.find({}, {"_id": 0}).to_list(100000),
        "suppliers": await db.suppliers.find({}, {"_id": 0}).to_list(10000),
        "sales": await db.sales.find({}, {"_id": 0}).to_list(100000),
        "nirs": await db.nirs.find({}, {"_id": 0}).to_list(10000),
        "users": await db.users.find({}, {"_id": 0, "password": 0}).to_list(1000)
    }
    return backup_data

@api_router.get("/backup/products-csv")
async def export_products_csv(user: dict = Depends(require_admin)):
    """Export all products as CSV for Excel"""
    products = await db.products.find({}, {"_id": 0}).to_list(100000)
    
    # Build CSV
    headers = ["Denumire", "Categorie", "Cod Bare", "Pret Achizitie", "Pret Vanzare", "TVA %", "Unitate", "Stoc", "Stoc Minim"]
    rows = []
    for p in products:
        rows.append([
            p.get("nume", ""),
            p.get("categorie", ""),
            p.get("cod_bare", "") or "",
            str(p.get("pret_achizitie", 0)),
            str(p.get("pret_vanzare", 0)),
            str(p.get("tva", 21)),
            p.get("unitate", "buc"),
            str(p.get("stoc", 0)),
            str(p.get("stoc_minim", 5))
        ])
    
    return {
        "headers": headers,
        "rows": rows,
        "total": len(rows)
    }

@api_router.get("/products/export/xls")
async def export_products_xls(user: dict = Depends(require_admin)):
    """Export all products as Excel XLS file"""
    products = await db.products.find({}, {"_id": 0}).to_list(100000)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Produse ANDREPAU"
    
    # Header styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="F59E0B", end_color="F59E0B", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = ["Nr.", "Denumire", "Categorie", "Cod Bare", "Preț Achiziție", "Preț Vânzare", "TVA %", "Unitate", "Stoc", "Stoc Minim"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Data rows
    for idx, p in enumerate(products, 1):
        row = [
            idx,
            p.get("nume", ""),
            p.get("categorie", ""),
            p.get("cod_bare", "") or "",
            p.get("pret_achizitie", 0),
            p.get("pret_vanzare", 0),
            p.get("tva", 21),
            p.get("unitate", "buc"),
            p.get("stoc", 0),
            p.get("stoc_minim", 5)
        ]
        for col, value in enumerate(row, 1):
            cell = ws.cell(row=idx + 1, column=col, value=value)
            cell.border = thin_border
            if col in [5, 6]:  # Price columns
                cell.number_format = '#,##0.00 "RON"'
            elif col in [9, 10]:  # Stock columns
                cell.number_format = '#,##0.00'
    
    # Column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 12
    ws.column_dimensions['J'].width = 12
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Generate filename with date
    filename = f"produse_andrepau_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# ==================== ANAF API PROXY ====================

# Cache for searched companies
companies_cache = {}

@api_router.post("/anaf/search-cui")
async def search_anaf_cui(data: dict, user: dict = Depends(get_current_user)):
    """Search company info from ANAF by CUI - with cache and OpenAPI.ro fallback"""
    import httpx
    
    cui = data.get("cui", "").strip()
    openapi_key = data.get("openapi_key", "").strip()  # Optional OpenAPI.ro key
    
    if not cui:
        raise HTTPException(status_code=400, detail="CUI este obligatoriu")
    
    # Remove 'RO' prefix if present
    cui_clean = cui.upper().replace("RO", "").strip()
    
    try:
        cui_int = int(cui_clean)
    except ValueError:
        raise HTTPException(status_code=400, detail="CUI invalid - trebuie să fie un număr")
    
    # Check cache first
    if cui_clean in companies_cache:
        logger.info(f"Returning cached company data for CUI {cui_clean}")
        return companies_cache[cui_clean]
    
    # Check database for previously saved company (local cache)
    existing = await db.companies_cache.find_one({"cui": cui_clean})
    if existing:
        cached_data = {
            "cui": existing.get("cui"),
            "denumire": existing.get("denumire"),
            "adresa": existing.get("adresa"),
            "nr_reg_com": existing.get("nr_reg_com"),
            "telefon": existing.get("telefon", ""),
            "cod_postal": existing.get("cod_postal", ""),
            "platitor_tva": existing.get("platitor_tva", False),
            "stare": existing.get("stare", ""),
            "localitate": existing.get("localitate", ""),
            "judet": existing.get("judet", ""),
            "from_cache": True
        }
        companies_cache[cui_clean] = cached_data
        return cached_data
    
    # Check romania_companies collection (from ONRC data.gov.ro)
    onrc_company = await db.romania_companies.find_one({"cui": cui_clean})
    if onrc_company:
        result_data = {
            "cui": onrc_company.get("cui"),
            "denumire": onrc_company.get("denumire"),
            "adresa": onrc_company.get("adresa"),
            "nr_reg_com": onrc_company.get("nr_reg_com"),
            "telefon": "",
            "cod_postal": onrc_company.get("cod_postal", ""),
            "platitor_tva": False,  # ONRC doesn't have this, would need ANAF
            "stare": "ACTIV",
            "localitate": onrc_company.get("localitate", ""),
            "judet": onrc_company.get("judet", ""),
            "source": "onrc",
            "from_cache": True
        }
        companies_cache[cui_clean] = result_data
        return result_data
    
    # Get current date
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    
    company_data = None
    
    # Try OpenAPI.ro first if key provided
    if openapi_key:
        try:
            async with httpx.AsyncClient(timeout=10.0) as client:
                response = await client.get(
                    f"https://api.openapi.ro/api/companies/{cui_int}",
                    headers={"x-api-key": openapi_key}
                )
                
                if response.status_code == 200:
                    result = response.json()
                    company_data = {
                        "cui": cui_clean,
                        "denumire": result.get("denumire", ""),
                        "adresa": result.get("adresa", ""),
                        "nr_reg_com": result.get("numar_reg_com", ""),
                        "telefon": result.get("telefon", ""),
                        "cod_postal": result.get("cod_postal", ""),
                        "platitor_tva": result.get("tva", False),
                        "stare": result.get("stare", ""),
                        "localitate": result.get("localitate", ""),
                        "judet": result.get("judet", ""),
                        "source": "openapi.ro"
                    }
        except Exception as e:
            logger.warning(f"OpenAPI.ro error: {str(e)}")
    
    # Try ANAF if OpenAPI didn't work
    if not company_data:
        anaf_url = "https://webservicesp.anaf.ro/PlatitorTvaRest/api/v8/ws/tva"
        payload = [{"cui": cui_int, "data": today}]
        
        try:
            async with httpx.AsyncClient(timeout=15.0) as client:
                response = await client.post(
                    anaf_url,
                    json=payload,
                    headers={
                        "Content-Type": "application/json",
                        "User-Agent": "ANDREPAU-POS/1.0"
                    }
                )
                
                # Check for HTML redirect (ANAF blocks some requests)
                if "<!DOCTYPE" in response.text or "<html" in response.text.lower():
                    logger.warning("ANAF returned HTML redirect - service may be blocking this IP")
                elif response.status_code == 200:
                    result = response.json()
                    
                    if result.get("found") and len(result["found"]) > 0:
                        company = result["found"][0]
                        date_generale = company.get("date_generale", {})
                        adresa_sediu = company.get("adresa_sediu_social", {})
                        inregistrare_tva = company.get("inregistrare_scop_Tva", {})
                        
                        # Build full address
                        address_parts = []
                        if adresa_sediu.get("sdenumire_Strada"):
                            street = adresa_sediu.get("sdenumire_Strada", "")
                            nr = adresa_sediu.get("snumar_Strada", "")
                            if nr:
                                address_parts.append(f"{street} {nr}")
                            else:
                                address_parts.append(street)
                        if adresa_sediu.get("sdenumire_Localitate"):
                            address_parts.append(adresa_sediu.get("sdenumire_Localitate"))
                        if adresa_sediu.get("sdenumire_Judet"):
                            address_parts.append(f"Jud. {adresa_sediu.get('sdenumire_Judet')}")
                        
                        full_address = ", ".join(address_parts) if address_parts else date_generale.get("adresa", "")
                        
                        company_data = {
                            "cui": cui_clean,
                            "denumire": date_generale.get("denumire", ""),
                            "adresa": full_address,
                            "nr_reg_com": date_generale.get("nrRegCom", ""),
                            "telefon": date_generale.get("telefon", ""),
                            "cod_postal": adresa_sediu.get("scod_Postal", "") or date_generale.get("codPostal", ""),
                            "platitor_tva": inregistrare_tva.get("scpTVA", False),
                            "stare": date_generale.get("stare_inregistrare", ""),
                            "localitate": adresa_sediu.get("sdenumire_Localitate", ""),
                            "judet": adresa_sediu.get("sdenumire_Judet", ""),
                            "source": "anaf"
                        }
                        
        except httpx.TimeoutException:
            logger.warning("ANAF timeout")
        except httpx.RequestError as e:
            logger.warning(f"ANAF request error: {str(e)}")
    
    if company_data:
        # Save to database cache
        await db.companies_cache.update_one(
            {"cui": cui_clean},
            {"$set": {**company_data, "updated_at": datetime.now(timezone.utc)}},
            upsert=True
        )
        # Save to memory cache
        companies_cache[cui_clean] = company_data
        return company_data
    
    raise HTTPException(
        status_code=503, 
        detail="Serviciul ANAF nu este disponibil momentan. Completați datele manual sau încercați din OpenAPI.ro (openapi.ro - 100 căutări gratuite/lună)."
    )

@api_router.post("/companies/save")
async def save_company_manually(data: dict, user: dict = Depends(get_current_user)):
    """Save a company manually to the local cache"""
    cui = data.get("cui", "").strip().upper().replace("RO", "")
    if not cui:
        raise HTTPException(status_code=400, detail="CUI este obligatoriu")
    
    company_data = {
        "cui": cui,
        "denumire": data.get("denumire", ""),
        "adresa": data.get("adresa", ""),
        "nr_reg_com": data.get("nr_reg_com", ""),
        "telefon": data.get("telefon", ""),
        "cod_postal": data.get("cod_postal", ""),
        "platitor_tva": data.get("platitor_tva", False),
        "stare": data.get("stare", "ACTIV"),
        "localitate": data.get("localitate", ""),
        "judet": data.get("judet", ""),
        "source": "manual",
        "updated_at": datetime.now(timezone.utc)
    }
    
    await db.companies_cache.update_one(
        {"cui": cui},
        {"$set": company_data},
        upsert=True
    )
    
    companies_cache[cui] = company_data
    return {"message": "Firmă salvată cu succes", "company": company_data}

@api_router.get("/companies/cached")
async def get_cached_companies(user: dict = Depends(get_current_user)):
    """Get all cached companies"""
    companies = await db.companies_cache.find({}, {"_id": 0}).to_list(1000)
    return {"companies": companies, "total": len(companies)}

# ==================== SEED DATA ====================

@api_router.post("/seed")
async def seed_database():
    # Check if already seeded
    existing_admin = await db.users.find_one({"username": "admin"})
    if existing_admin:
        return {"message": "Database deja populată"}
    
    # Create admin user
    admin_doc = {
        "id": str(uuid.uuid4()),
        "username": "admin",
        "password": hash_password("admin123"),
        "full_name": "Administrator",
        "role": "admin",
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.users.insert_one(admin_doc)
    
    # Create casier user
    casier_doc = {
        "id": str(uuid.uuid4()),
        "username": "casier",
        "password": hash_password("casier123"),
        "full_name": "Casier Principal",
        "role": "casier",
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.users.insert_one(casier_doc)
    
    # Create suppliers
    suppliers = [
        {"id": str(uuid.uuid4()), "nume": "Dedeman S.R.L.", "telefon": "0721000001", "email": "contact@dedeman.ro", "adresa": "București", "created_at": datetime.now(timezone.utc).isoformat()},
        {"id": str(uuid.uuid4()), "nume": "Hornbach România", "telefon": "0721000002", "email": "contact@hornbach.ro", "adresa": "Cluj-Napoca", "created_at": datetime.now(timezone.utc).isoformat()},
        {"id": str(uuid.uuid4()), "nume": "Leroy Merlin", "telefon": "0721000003", "email": "contact@leroymerlin.ro", "adresa": "Timișoara", "created_at": datetime.now(timezone.utc).isoformat()},
    ]
    await db.suppliers.insert_many(suppliers)
    
    # Create products
    categories = ["Materiale Construcții", "Scule Electrice", "Scule Manuale", "Feronerie", "Instalații Sanitare", "Electrice", "Vopsele", "Consumabile"]
    units = ["buc", "sac", "kg", "metru", "litru", "rola"]
    
    products = [
        # Materiale Construcții
        {"id": str(uuid.uuid4()), "nume": "Ciment Portland 40kg", "categorie": "Materiale Construcții", "furnizor_id": suppliers[0]["id"], "cod_bare": "5941234000001", "pret_achizitie": 28.0, "pret_vanzare": 35.0, "tva": 21.0, "unitate": "sac", "stoc": 150, "stoc_minim": 20, "descriere": "Ciment de înaltă calitate", "created_at": datetime.now(timezone.utc).isoformat()},
        {"id": str(uuid.uuid4()), "nume": "Var hidratat 20kg", "categorie": "Materiale Construcții", "furnizor_id": suppliers[0]["id"], "cod_bare": "5941234000002", "pret_achizitie": 12.0, "pret_vanzare": 18.0, "tva": 21.0, "unitate": "sac", "stoc": 80, "stoc_minim": 15, "descriere": "Var pentru construcții", "created_at": datetime.now(timezone.utc).isoformat()},
        {"id": str(uuid.uuid4()), "nume": "Nisip fin 40kg", "categorie": "Materiale Construcții", "furnizor_id": suppliers[0]["id"], "cod_bare": "5941234000003", "pret_achizitie": 8.0, "pret_vanzare": 12.0, "tva": 21.0, "unitate": "sac", "stoc": 200, "stoc_minim": 30, "descriere": "Nisip pentru tencuială", "created_at": datetime.now(timezone.utc).isoformat()},
        {"id": str(uuid.uuid4()), "nume": "BCA 60x20x25", "categorie": "Materiale Construcții", "furnizor_id": suppliers[1]["id"], "cod_bare": "5941234000004", "pret_achizitie": 6.5, "pret_vanzare": 9.0, "tva": 21.0, "unitate": "buc", "stoc": 500, "stoc_minim": 50, "descriere": "Blocuri BCA", "created_at": datetime.now(timezone.utc).isoformat()},
        
        # Scule Electrice
        {"id": str(uuid.uuid4()), "nume": "Bormaşină Bosch 750W", "categorie": "Scule Electrice", "furnizor_id": suppliers[1]["id"], "cod_bare": "5941234000010", "pret_achizitie": 180.0, "pret_vanzare": 250.0, "tva": 21.0, "unitate": "buc", "stoc": 15, "stoc_minim": 3, "descriere": "Bormaşină profesională", "created_at": datetime.now(timezone.utc).isoformat()},
        {"id": str(uuid.uuid4()), "nume": "Flex 125mm 1200W", "categorie": "Scule Electrice", "furnizor_id": suppliers[1]["id"], "cod_bare": "5941234000011", "pret_achizitie": 120.0, "pret_vanzare": 170.0, "tva": 21.0, "unitate": "buc", "stoc": 12, "stoc_minim": 2, "descriere": "Polizor unghiular", "created_at": datetime.now(timezone.utc).isoformat()},
        {"id": str(uuid.uuid4()), "nume": "Șurubelniță electrică", "categorie": "Scule Electrice", "furnizor_id": suppliers[2]["id"], "cod_bare": "5941234000012", "pret_achizitie": 85.0, "pret_vanzare": 120.0, "tva": 21.0, "unitate": "buc", "stoc": 20, "stoc_minim": 5, "descriere": "Autofiletantă cu acumulator", "created_at": datetime.now(timezone.utc).isoformat()},
        
        # Scule Manuale
        {"id": str(uuid.uuid4()), "nume": "Ciocan 500g", "categorie": "Scule Manuale", "furnizor_id": suppliers[0]["id"], "cod_bare": "5941234000020", "pret_achizitie": 25.0, "pret_vanzare": 38.0, "tva": 21.0, "unitate": "buc", "stoc": 40, "stoc_minim": 10, "descriere": "Ciocan cu mâner fibră", "created_at": datetime.now(timezone.utc).isoformat()},
        {"id": str(uuid.uuid4()), "nume": "Set șurubelnițe 6buc", "categorie": "Scule Manuale", "furnizor_id": suppliers[2]["id"], "cod_bare": "5941234000021", "pret_achizitie": 35.0, "pret_vanzare": 55.0, "tva": 21.0, "unitate": "buc", "stoc": 25, "stoc_minim": 5, "descriere": "Set complet șurubelnițe", "created_at": datetime.now(timezone.utc).isoformat()},
        {"id": str(uuid.uuid4()), "nume": "Cheie reglabilă 300mm", "categorie": "Scule Manuale", "furnizor_id": suppliers[1]["id"], "cod_bare": "5941234000022", "pret_achizitie": 45.0, "pret_vanzare": 65.0, "tva": 21.0, "unitate": "buc", "stoc": 18, "stoc_minim": 4, "descriere": "Cheie franceză", "created_at": datetime.now(timezone.utc).isoformat()},
        
        # Feronerie
        {"id": str(uuid.uuid4()), "nume": "Șuruburi 4x40 100buc", "categorie": "Feronerie", "furnizor_id": suppliers[0]["id"], "cod_bare": "5941234000030", "pret_achizitie": 8.0, "pret_vanzare": 14.0, "tva": 21.0, "unitate": "buc", "stoc": 100, "stoc_minim": 20, "descriere": "Șuruburi autofiletante", "created_at": datetime.now(timezone.utc).isoformat()},
        {"id": str(uuid.uuid4()), "nume": "Diblu plastic 8mm 100buc", "categorie": "Feronerie", "furnizor_id": suppliers[0]["id"], "cod_bare": "5941234000031", "pret_achizitie": 6.0, "pret_vanzare": 10.0, "tva": 21.0, "unitate": "buc", "stoc": 150, "stoc_minim": 30, "descriere": "Dibluri pentru BCA", "created_at": datetime.now(timezone.utc).isoformat()},
        {"id": str(uuid.uuid4()), "nume": "Balama ușă 100mm", "categorie": "Feronerie", "furnizor_id": suppliers[2]["id"], "cod_bare": "5941234000032", "pret_achizitie": 12.0, "pret_vanzare": 18.0, "tva": 21.0, "unitate": "buc", "stoc": 60, "stoc_minim": 10, "descriere": "Balama din oțel", "created_at": datetime.now(timezone.utc).isoformat()},
        
        # Instalații Sanitare
        {"id": str(uuid.uuid4()), "nume": "Țeavă PPR 25mm", "categorie": "Instalații Sanitare", "furnizor_id": suppliers[1]["id"], "cod_bare": "5941234000040", "pret_achizitie": 3.5, "pret_vanzare": 5.5, "tva": 21.0, "unitate": "metru", "stoc": 500, "stoc_minim": 100, "descriere": "Țeavă apă caldă/rece", "created_at": datetime.now(timezone.utc).isoformat()},
        {"id": str(uuid.uuid4()), "nume": "Robinet colț 1/2", "categorie": "Instalații Sanitare", "furnizor_id": suppliers[1]["id"], "cod_bare": "5941234000041", "pret_achizitie": 18.0, "pret_vanzare": 28.0, "tva": 21.0, "unitate": "buc", "stoc": 40, "stoc_minim": 10, "descriere": "Robinet colțar", "created_at": datetime.now(timezone.utc).isoformat()},
        {"id": str(uuid.uuid4()), "nume": "Sifon chiuvetă", "categorie": "Instalații Sanitare", "furnizor_id": suppliers[2]["id"], "cod_bare": "5941234000042", "pret_achizitie": 15.0, "pret_vanzare": 25.0, "tva": 21.0, "unitate": "buc", "stoc": 35, "stoc_minim": 8, "descriere": "Sifon plastic", "created_at": datetime.now(timezone.utc).isoformat()},
        
        # Electrice
        {"id": str(uuid.uuid4()), "nume": "Cablu electric 2.5mm", "categorie": "Electrice", "furnizor_id": suppliers[0]["id"], "cod_bare": "5941234000050", "pret_achizitie": 4.0, "pret_vanzare": 6.5, "tva": 21.0, "unitate": "metru", "stoc": 800, "stoc_minim": 200, "descriere": "Cablu FY cupru", "created_at": datetime.now(timezone.utc).isoformat()},
        {"id": str(uuid.uuid4()), "nume": "Priză simplă", "categorie": "Electrice", "furnizor_id": suppliers[2]["id"], "cod_bare": "5941234000051", "pret_achizitie": 8.0, "pret_vanzare": 14.0, "tva": 21.0, "unitate": "buc", "stoc": 100, "stoc_minim": 25, "descriere": "Priză cu împământare", "created_at": datetime.now(timezone.utc).isoformat()},
        {"id": str(uuid.uuid4()), "nume": "Întrerupător simplu", "categorie": "Electrice", "furnizor_id": suppliers[2]["id"], "cod_bare": "5941234000052", "pret_achizitie": 7.0, "pret_vanzare": 12.0, "tva": 21.0, "unitate": "buc", "stoc": 90, "stoc_minim": 20, "descriere": "Întrerupător alb", "created_at": datetime.now(timezone.utc).isoformat()},
        
        # Vopsele
        {"id": str(uuid.uuid4()), "nume": "Vopsea lavabilă albă 15L", "categorie": "Vopsele", "furnizor_id": suppliers[1]["id"], "cod_bare": "5941234000060", "pret_achizitie": 85.0, "pret_vanzare": 120.0, "tva": 21.0, "unitate": "buc", "stoc": 30, "stoc_minim": 5, "descriere": "Vopsea interior", "created_at": datetime.now(timezone.utc).isoformat()},
        {"id": str(uuid.uuid4()), "nume": "Grund alb 10L", "categorie": "Vopsele", "furnizor_id": suppliers[1]["id"], "cod_bare": "5941234000061", "pret_achizitie": 55.0, "pret_vanzare": 80.0, "tva": 21.0, "unitate": "buc", "stoc": 25, "stoc_minim": 5, "descriere": "Grund acrilic", "created_at": datetime.now(timezone.utc).isoformat()},
        {"id": str(uuid.uuid4()), "nume": "Vopsea email albastru 0.75L", "categorie": "Vopsele", "furnizor_id": suppliers[0]["id"], "cod_bare": "5941234000062", "pret_achizitie": 28.0, "pret_vanzare": 42.0, "tva": 21.0, "unitate": "buc", "stoc": 45, "stoc_minim": 10, "descriere": "Email pentru metal", "created_at": datetime.now(timezone.utc).isoformat()},
        
        # Consumabile
        {"id": str(uuid.uuid4()), "nume": "Bandă izolatoare", "categorie": "Consumabile", "furnizor_id": suppliers[0]["id"], "cod_bare": "5941234000070", "pret_achizitie": 3.0, "pret_vanzare": 5.0, "tva": 21.0, "unitate": "buc", "stoc": 200, "stoc_minim": 50, "descriere": "Bandă PVC", "created_at": datetime.now(timezone.utc).isoformat()},
        {"id": str(uuid.uuid4()), "nume": "Silicon transparent 280ml", "categorie": "Consumabile", "furnizor_id": suppliers[2]["id"], "cod_bare": "5941234000071", "pret_achizitie": 12.0, "pret_vanzare": 18.0, "tva": 21.0, "unitate": "buc", "stoc": 80, "stoc_minim": 20, "descriere": "Silicon universal", "created_at": datetime.now(timezone.utc).isoformat()},
        {"id": str(uuid.uuid4()), "nume": "Disc flex 125mm metal", "categorie": "Consumabile", "furnizor_id": suppliers[1]["id"], "cod_bare": "5941234000072", "pret_achizitie": 4.0, "pret_vanzare": 7.0, "tva": 21.0, "unitate": "buc", "stoc": 150, "stoc_minim": 30, "descriere": "Disc abraziv", "created_at": datetime.now(timezone.utc).isoformat()},
        
        # Low stock products for testing alerts
        {"id": str(uuid.uuid4()), "nume": "Furtun grădină 20m", "categorie": "Consumabile", "furnizor_id": suppliers[0]["id"], "cod_bare": "5941234000080", "pret_achizitie": 45.0, "pret_vanzare": 65.0, "tva": 21.0, "unitate": "buc", "stoc": 3, "stoc_minim": 5, "descriere": "Furtun cu armătură", "created_at": datetime.now(timezone.utc).isoformat()},
        {"id": str(uuid.uuid4()), "nume": "Mănuși protecție", "categorie": "Consumabile", "furnizor_id": suppliers[2]["id"], "cod_bare": "5941234000081", "pret_achizitie": 8.0, "pret_vanzare": 15.0, "tva": 21.0, "unitate": "buc", "stoc": 2, "stoc_minim": 10, "descriere": "Mănuși latex", "created_at": datetime.now(timezone.utc).isoformat()},
    ]
    await db.products.insert_many(products)
    
    return {"message": "Database populată cu succes", "users": 2, "suppliers": 3, "products": len(products)}

# ==================== ROOT ====================

@api_router.get("/")
async def root():
    return {"message": "ANDREPAU POS API", "version": "1.0.0"}

# Include the router in the main app
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
